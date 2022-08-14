# python包
import qdarkstyle
import sys
import openpyxl
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (QApplication, QMainWindow, QAction,
                             QSplitter, QInputDialog, QFileDialog,
                             QMessageBox)

# 自建包
import CodeTools.loggingTool as logTool
from CodeTools.loggingTool import PotsLogger
from UIWidgets.startForm import PotsStartFormBasic
from UIWidgets.IOForm import PotsIOFormBasic
from UIWidgets.delegates import PotsStartFormViewDelegate, PotsIoFormViewDelegate
from UIWidgets.treeWindow import PotsTree
from UIWidgets.mdiWindow import PotsMdiWindow
from DataManage.DBManage import PotsDataVessel
from DataManage.ObjManage import PotsGlobalVessel, PotsVessel, PotsConfirmForms
from UIWidgets.childWindow import PotsChildWindow_manager
from UIWidgets.coverWindow import PotsCoverWindow

# 类
class PotsLogsVar():
    def __init__(self):
        # 日志共用变量初始化
        self.path_exe = logTool.PotsLogFolder.path_exe
        self.global_tlf = logTool.PotsLogFolder.topFolder_log
        self.global_lf = logTool.PotsLogFolder.folder_log
        self.global_fb = logTool.PotsLogFile.flie_bessineseLog
        self.global_fv = logTool.PotsLogFile.file_varLog


class PotsCelUI(QMainWindow):
    def __init__(self):
        super(PotsCelUI, self).__init__(parent=None)
        self.setWindowTitle('IO清册工具')
        self.resize(600, 900)
        self.logger = PotsLogger('PotsCelUI')
        self.logger.getFuncName('入口文件进入初始化')

        self.initVessel()
        self.initLogic()
        self.initWidgets()
        self.initTopologica()

    def closeEvent(self, event):
        self.logger.getFuncName('closeEvent')
        self.logger.normalInfo('创建标识符字符串')
        str_unSavedProjectTitle = ''
        for vessel_project in self.allVessel:
            if vessel_project[9] == False:
                str_unSavedProjectTitle += str(vessel_project[0])+'\n'
            else:
                continue

        if str_unSavedProjectTitle != '':
            result = QMessageBox.question(self, '提示', '以下项目未保存，确认关闭？ \n'+str_unSavedProjectTitle,
                                     QMessageBox.Yes|QMessageBox.No, QMessageBox.No)
            if result == QMessageBox.Yes:
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()

    # 初始化容器
    def initVessel(self):
        self.logger.getFuncName('initVessel')
        # 容器及对象初始化
        self.logger.normalInfo('初始化容器')
        self.obj_allVessel = PotsGlobalVessel()
        self.allVessel = self.obj_allVessel.globalVessel
        self.dataVessel = PotsDataVessel()
        self.treeObject = PotsTree(self, self.allVessel)
        self.mdiObject = PotsMdiWindow(self, self.treeObject)

    # 获取并初始化自定义的控件
    def initWidgets(self):
        self.logger.getFuncName('initWidgets')
        self.logger.normalInfo('获取并初始化自定义的控件')

        menuBar = self.menuBar()
        fileMenu = menuBar.addMenu('文件')
        newMenu = fileMenu.addMenu('新建')
        newMenu.addAction(self.aFileNewCurrent)
        newMenu.addAction(self.aFileNewHighDensity)
        fileMenu.addAction(self.aFileOpen)
        fileMenu.addAction(self.aFileSave)
        saveAsMenu = fileMenu.addMenu('另存为')
        # saveAsMenu.addAction(self.aFileSaveAsDB)
        saveAsMenu.addAction(self.aFileSaveAsCel)
        editMenu = menuBar.addMenu('编辑')
        addCartonsMenu = editMenu.addMenu('添加自定义卡件')
        addCartonsMenu.addAction(self.aEditAddCarton)
        formatMenu = menuBar.addMenu('格式')
        helpMenu = menuBar.addMenu('帮助')

        toolBar = self.addToolBar('')
        toolBar.addAction(self.aSheetCreate)
        toolBar.addAction(self.aAnalize)
        toolBar.addAction(self.aCheck)
        toolBar.addAction(self.aMachinCel)

    # 布局控件
    def initTopologica(self):
        self.logger.getFuncName('initTopologica')
        self.logger.normalInfo('布局控件')
        self.hSpilter = QSplitter(self)
        self.hSpilter.addWidget(self.treeObject.tree)
        self.hSpilter.addWidget(self.mdiObject.mdi)
        self.setCentralWidget(self.hSpilter)

    # 初始化逻辑代码并向业务逻辑代码添加控件对象
    def initLogic(self):
        self.logger.getFuncName('initLogic')

        # 菜单动作
        # 动作——新建通用机柜
        self.logger.normalInfo('新建通用机柜')
        self.aFileNewCurrent = QAction('通用', self)
        self.aFileNewCurrent.setToolTip('新建通用机柜项目')  # 添加提示消息
        self.aFileNewCurrent.triggered.connect(self.FileNewCurrent)  # 将动作绑定至功能函数
        # 动作——新建高密度机柜
        self.logger.normalInfo('新建高密度机柜')
        self.aFileNewHighDensity = QAction('高密度', self)
        self.aFileNewHighDensity.setToolTip('新建高密度机柜项目')  # 添加提示消息
        self.aFileNewHighDensity.triggered.connect(self.FileNewHighDensity)  # 将动作绑定至功能函数
        # 动作——打开文件
        self.logger.normalInfo('打开文件')
        self.aFileOpen = QAction('打开', self)
        self.aFileOpen.setToolTip('打开一个文件')  # 添加提示消息
        self.aFileOpen.triggered.connect(self.FileOpen)  # 将动作绑定至功能函数
        # 动作——保存
        self.logger.normalInfo('保存')
        self.aFileSave = QAction('保存', self)
        self.aFileSave.setToolTip('保存文件')  # 添加提示消息
        self.aFileSave.triggered.connect(self.FileSave)  # 将动作绑定至功能函数
        """
        # 动作——另存为(DB)
        self.aFileSaveAsDB = QAction('另存为DB文件', self)
        self.aFileSaveAsDB.setToolTip('另存为EXCEL文件')  # 添加提示消息
        self.aFileSaveAsDB.triggered.connect(...)  # 将动作绑定至功能函数
        """
        # 动作——另存为(cel)
        self.logger.normalInfo('另存为(cel)')
        self.aFileSaveAsCel = QAction('另存为Excel', self)
        self.aFileSaveAsCel.setToolTip('另存为')  # 添加提示消息
        self.aFileSaveAsCel.triggered.connect(self.FileSaveAsCel)  # 将动作绑定至功能函数
        # 动作——添加自定义卡件
        self.logger.normalInfo('添加自定义卡件')
        self.aEditAddCarton = QAction('自定义卡件', self)
        self.aEditAddCarton.setToolTip('添加一个自定义卡件')  # 添加提示消息
        self.aEditAddCarton.triggered.connect(self.EditAddCarton)  # 将动作绑定至功能函数

        # 工具栏动作
        # 动作——清册生成
        self.logger.normalInfo('动作——清册生成')
        self.aSheetCreate = QAction('清册生成', self)
        self.aSheetCreate.setToolTip('按导入数据生成清册')
        self.aSheetCreate.triggered.connect(self.IOSheetCreate)
        # 动作——统计分析
        self.logger.normalInfo('动作——统计分析')
        self.aAnalize = QAction('统计分析', self)
        self.aAnalize.setToolTip('分析表格数据')
        self.aAnalize.triggered.connect(self.Analize)
        # 动作——检查
        self.logger.normalInfo('动作——检查')
        self.aCheck = QAction('检查', self)
        self.aCheck.setToolTip('检查数据错误')
        self.aCheck.triggered.connect(self.Check)
        # 动作——设备表
        self.logger.normalInfo('动作——设备表')
        self.aMachinCel = QAction('设备表', self)
        self.aMachinCel.setToolTip('按数据生成设备表格')
        self.aMachinCel.triggered.connect(self.MachinCel)

        """
                # 动作——撤销编辑
                self.aEditUndo = QAction('撤销', self)
                self.aEditUndo.triggered.connect(self.txtEditor.undo)
                # 动作——恢复编辑
                self.aEditRedo = QAction('恢复', self)
                self.aEditRedo.triggered.connect(self.txtEditor.redo)
            """

    def FileNewCurrent(self):
        self.logger.getFuncName('FileNewCurrent')
        try:
            name = None
            self.logger.normalInfo('弹出新建对话框')
            text, ok = QInputDialog.getText(self, '项目命名', '输入名称')
            self.logger.varInfo('项目名称->'+str(text))
            if ok:
                name = str(text)
        except Exception as e:
            self.logger.warringInfo(str(e))
            dlg = QMessageBox(self)
            dlg.setIcon(QMessageBox.Critical)
            dlg.setText(str(e))
            dlg.show()
            return 0

        if name is None or name == '':
            self.logger.warringInfo('项目新建对话框未获得项目名')
            dlg = QMessageBox(self)
            dlg.setIcon(QMessageBox.Critical)
            dlg.setText('项目新建对话框未获得项目名')
            dlg.show()
            return 0

        self.logger.normalInfo('创建掩盖窗体')
        self.coverWindow = PotsCoverWindow(self)
        self.logger.normalInfo('创建完毕')

        self.logger.normalInfo('设置进度为12')
        self.coverWindow.setValue(12)

        self.logger.normalInfo('新建项目容器')
        self.projectVessel = PotsVessel(name)
        self.logger.normalInfo('添加项目容器到总容器')
        self.allVessel.append(self.projectVessel.project)

        self.logger.normalInfo('设置进度为20')
        self.coverWindow.setValue(20)

        self.logger.normalInfo('新建树结构')
        self.treeObject.getInfo('通用', self.projectVessel, self.mdiObject)
        self.treeObject.addProject(name)

        self.logger.normalInfo('设置进度为25')
        self.coverWindow.setValue(25)

        self.logger.normalInfo('新建模件表->通用')
        startModelCreator = PotsStartFormBasic(self, '模板', name, '通用', self.projectVessel)

        self.logger.normalInfo('设置进度为28')
        self.coverWindow.setValue(28)

        self.logger.normalInfo('新建模件表代理')
        delegate_startForm = PotsStartFormViewDelegate(self)
        delegate_startForm.getModel(self.projectVessel.project[4][0][0], self.projectVessel)
        delegate_startForm.getChoosesAndIndex('通用', self.dataVessel)
        self.logger.normalInfo('为模件表添加代理')
        startModelCreator.getDelegate(delegate_startForm)

        self.logger.normalInfo('设置进度为30')
        self.coverWindow.setValue(30)

        self.logger.normalInfo('新建IO清册表单')
        IoModelCreator = PotsIOFormBasic(self, '清册', name, '通用', self.projectVessel)

        self.logger.normalInfo('设置进度为50')
        self.coverWindow.setValue(50)

        self.logger.normalInfo('进入循环，为每个IO表单赋值')
        value_add = 0
        for IoModleList in self.projectVessel.project[5]:
            value_add += 1
            self.logger.varInfo('项目容器中的IO表单列表->'+str(IoModleList))
            delegate_IoForm = PotsIoFormViewDelegate()
            delegate_IoForm.getModel(IoModleList[0], self.projectVessel)
            delegate_IoForm.getChoosesAndIndex(self.dataVessel)
            IoModelCreator.getDelegate(delegate_IoForm)
            self.coverWindow.setValue(value_add + 50)

        self.logger.normalInfo('关闭掩盖窗体')
        self.coverWindow.timer.stop()
        self.coverWindow.close()

        self.logger.varInfo('FileNewCurrent')
        self.logger.varInfo('此时的容器内容->\n'+str(self.projectVessel.project))
        self.logger.varInfo('此时的总容器内容->\n'+str(self.allVessel))

        self.logger.normalInfo('新建通用完毕，牛逼')

    def FileNewHighDensity(self):
        self.logger.getFuncName('FileNewHighDensity')
        try:
            name = None
            self.logger.normalInfo('弹出新建对话框')
            text, ok = QInputDialog.getText(self, '项目命名', '输入名称')
            self.logger.varInfo('项目名称->'+str(text))
            if ok:
                name = str(text)
        except Exception as e:
            self.logger.warringInfo(str(e))
            dlg = QMessageBox(self)
            dlg.setIcon(QMessageBox.Critical)
            dlg.setText(str(e))
            dlg.show()
            return 0

        if name is None or name == '':
            self.logger.warringInfo('项目新建对话框未获得项目名')
            dlg = QMessageBox(self)
            dlg.setIcon(QMessageBox.Critical)
            dlg.setText('项目新建对话框未获得项目名')
            dlg.show()
            return 0

        self.logger.normalInfo('创建掩盖窗体')
        self.coverWindow = PotsCoverWindow(self)

        self.logger.normalInfo('设置进度为12')
        self.coverWindow.setValue(12)

        self.logger.normalInfo('新建项目容器')
        self.projectVessel = PotsVessel(name)
        self.logger.normalInfo('添加项目容器到总容器')
        self.allVessel.append(self.projectVessel.project)

        self.logger.normalInfo('设置进度为20')
        self.coverWindow.setValue(20)

        self.logger.normalInfo('新建树结构')
        self.treeObject.getInfo('高密度', self.projectVessel, self.mdiObject)
        self.treeObject.addProject(name)

        self.logger.normalInfo('设置进度为25')
        self.coverWindow.setValue(25)

        self.logger.normalInfo('新建模件表->高密度')
        startModelCreator = PotsStartFormBasic(self, '模板', name, '高密度', self.projectVessel)

        self.logger.normalInfo('设置进度为28')
        self.coverWindow.setValue(28)

        self.logger.normalInfo('新建模件表代理')
        delegate_startForm = PotsStartFormViewDelegate(self)
        delegate_startForm.getModel(self.projectVessel.project[4][0][0], self.projectVessel)
        delegate_startForm.getChoosesAndIndex('高密度', self.dataVessel)
        self.logger.normalInfo('为模件表添加代理')
        startModelCreator.getDelegate(delegate_startForm)

        self.logger.normalInfo('设置进度为30')
        self.coverWindow.setValue(30)

        self.logger.normalInfo('新建IO清册表单')
        IoModelCreator = PotsIOFormBasic(self, '清册', name, '高密度', self.projectVessel)

        self.logger.normalInfo('设置进度为50')
        self.coverWindow.setValue(50)

        self.logger.normalInfo('进入循环，为每个IO表单赋值')
        value_add = 0
        for IoModleList in self.projectVessel.project[5]:
            value_add += 1

            self.logger.varInfo('项目容器中的IO表单列表->' + str(IoModleList))
            delegate_IoForm = PotsIoFormViewDelegate()
            delegate_IoForm.getModel(IoModleList[0], self.projectVessel)
            delegate_IoForm.getChoosesAndIndex(self.dataVessel)
            IoModelCreator.getDelegate(delegate_IoForm)

            self.logger.normalInfo('设置进度为' + str(value_add + 50))
            self.coverWindow.setValue(value_add + 50)

        self.logger.normalInfo('关闭掩盖窗体')
        self.coverWindow.timer.stop()
        self.coverWindow.close()

        self.logger.varInfo('FileNewHighDensity')
        self.logger.varInfo('此时的容器内容->\n' + str(self.projectVessel.project))
        self.logger.varInfo('此时的总容器内容->\n' + str(self.allVessel))

        self.logger.normalInfo('新建高密度完毕，牛逼')

    def FileOpen(self):
        self.logger.getFuncName('FileOpen')
        try:
            self.logger.normalInfo('弹出打开文件对话框')
            path, __ = QFileDialog.getOpenFileName(self, '打开文件', '', '文件(*.xlsx)')
            if path is not None and path != '':
                self.logger.normalInfo('打开对话框获取到了路径')
                wb = openpyxl.load_workbook(path)
                self.logger.varInfo('打开操作获取到的路径值为->'+str(path))
            else:
                self.logger.warringInfo('对话框未获取到路径')
                return 0
        except Exception as e:
            self.logger.warringInfo(str(e))
            dlg = QMessageBox(self)
            dlg.setIcon(QMessageBox.Critical)
            dlg.setText(str(e))
            dlg.show()
            return 0
        else:
            name = '文件不规范'
            self.logger.normalInfo('进入循环，遍历文件获取工作簿，并获取所打开文件的项目名称')
            for sheet in wb:
                sheetName = sheet.title
                if '模件' in sheetName:
                    self.logger.normalInfo('成功匹配到含模件字样的工作簿')
                    name = sheetName[5:]
                    if '(H)' in sheetName:
                        cabinet = '高密度'
                    elif '(C)' in sheetName:
                        cabinet = '通用'
                    break

        if name == '文件不规范':
            self.logger.warringInfo('打开的文件里面没有含模件字样的工作簿')
            dlg = QMessageBox(self)
            dlg.setIcon(QMessageBox.Critical)
            dlg.setText(name)
            dlg.show()
            return 0

        self.logger.normalInfo('创建掩盖窗体')
        self.coverWindow = PotsCoverWindow(self)

        self.logger.normalInfo('建立对应名称的项目容器')
        self.projectVessel = PotsVessel(name)

        self.logger.normalInfo('设置进度为12')
        self.coverWindow.setValue(12)

        property_path = {
            'name': name
        }
        self.logger.normalInfo('向项目容器内添加路径')
        self.projectVessel.project[8][0][0] = path
        self.projectVessel.project[8][0][1] = property_path

        self.logger.normalInfo('设置进度为20')
        self.coverWindow.setValue(20)

        self.logger.normalInfo('添加项目容器到总容器')
        self.allVessel.append(self.projectVessel.project)
        self.logger.normalInfo('新建树结构')
        self.treeObject.getInfo(cabinet, self.projectVessel, self.mdiObject)
        self.treeObject.addProject(name)

        self.logger.normalInfo('设置进度为25')
        self.coverWindow.setValue(25)

        self.logger.normalInfo('新建模件表->高密度/通用')
        startModelCreator = PotsStartFormBasic(self, '模板', name, cabinet, self.projectVessel, True)

        self.logger.normalInfo('设置进度为28')
        self.coverWindow.setValue(28)

        self.logger.normalInfo('为模件表添加代理')
        delegate_startForm = PotsStartFormViewDelegate(self)
        delegate_startForm.getModel(self.projectVessel.project[4][0][0], self.projectVessel)
        delegate_startForm.getChoosesAndIndex(cabinet, self.dataVessel)
        startModelCreator.getDelegate(delegate_startForm)

        self.logger.normalInfo('设置进度为30')
        self.coverWindow.setValue(30)

        self.logger.normalInfo('新建IO清册表单')
        IoModelCreator = PotsIOFormBasic(self, '清册', name, cabinet, self.projectVessel, True)

        self.logger.normalInfo('设置进度为50')
        self.coverWindow.setValue(50)

        self.logger.normalInfo('进入循环，为每个IO表单赋值')
        value_add = 0
        for IoModleList in self.projectVessel.project[5]:
            value_add += 1
            self.logger.varInfo('项目容器中的IO表单列表->' + str(IoModleList))
            delegate_IoForm = PotsIoFormViewDelegate()
            delegate_IoForm.getModel(IoModleList[0], self.projectVessel)
            delegate_IoForm.getChoosesAndIndex(self.dataVessel)
            IoModelCreator.getDelegate(delegate_IoForm)

            index = self.projectVessel.project[5].index(IoModleList)
            self.projectVessel.project[7][index][1]['mergeList'] = self.projectVessel.project[5][index][1]['mergeList']
            mergeList = self.projectVessel.project[7][index][1]['mergeList']
            view = self.projectVessel.project[7][index][0]

            for coordinate_merge in mergeList:
                view.setSpan(coordinate_merge[0],
                                coordinate_merge[1],
                                coordinate_merge[2],
                                coordinate_merge[3])

            self.logger.normalInfo('设置进度为'+str(value_add+50))
            self.coverWindow.setValue(value_add+50)

        self.logger.normalInfo('关闭掩盖窗体')
        self.coverWindow.close()

        self.logger.varInfo('FileOpen')
        self.logger.varInfo('此时的容器内容->\n' + str(self.projectVessel.project))
        self.logger.varInfo('此时的总容器内容->\n' + str(self.allVessel))

        self.logger.normalInfo('打开项目完毕，牛逼')

    def FileSave(self, projectName=None):
        self.logger.getFuncName('FileSave')
        if projectName is False:
            self.logger.normalInfo('无项目名称传入时，获取当前窗口名称')
            __, name = self.mdiObject.getCurrentWidgets()
            if name is None:
                self.logger.warringInfo('当前无活动窗口')
                dlg = QMessageBox(self)
                dlg.setIcon(QMessageBox.Critical)
                dlg.setText('当前必须存在模板或清册窗口')
                dlg.show()
                return 0
            elif '清册' in name:
                projectName = name[4:]
            elif '模板' in name:
                projectName = name[3:]
            else:
                pass
            self.logger.normalInfo('匹配项目名称')

        self.logger.normalInfo('进入循环，遍历总容器内所有项目的路径')
        for project in self.allVessel:
            self.logger.varInfo('这是当前路径内容-> ' + str(project[8][0][0]))
            if projectName == project[0]:
                self.logger.normalInfo('根据名称匹配成功')
                if project[8][0][0] is None:
                    self.logger.normalInfo('发现项目容器内无路径,进入另存为操作')
                    self.FileSaveAsCel()
                    return 0
                else:
                    self.logger.normalInfo('成功获取到项目路径')
                    path = project[8][0][0]

                project[9] = True

                self.logger.normalInfo('创建掩盖窗体')
                self.coverWindow = PotsCoverWindow(self)

                self.logger.normalInfo('设置进度为3')
                self.coverWindow.setValue(3)

                self.logger.normalInfo('根据路径打开工作簿')
                wb = openpyxl.load_workbook(path)
                self.logger.normalInfo('遍历工作簿，并通过删除所有表单预处理')
                for sheet in wb:
                    name = sheet.title
                    del wb[name]

                self.logger.normalInfo('设置进度为20')
                self.coverWindow.setValue(20)

                # 模板页保存
                self.logger.normalInfo('保存模板页')
                self.logger.normalInfo('设置模件表单名称')
                if project[4][0][1]['cabinet'] == '高密度':
                    sheetName_startForm = '模件(H)' + projectName
                elif project[4][0][1]['cabinet'] == '通用':
                    sheetName_startForm = '模件(C)' + projectName
                self.logger.normalInfo('创建模件表单')
                sheet_startForm = wb.create_sheet(sheetName_startForm, 0)

                self.logger.normalInfo('获取项目容器内的模件表单及其行列数')
                model_startForm = project[4][0][0]
                rows_startForm = model_startForm.rowCount()
                cols_startForm = model_startForm.columnCount()
                self.logger.normalInfo('根据行列填入工作簿每个单元格的值')

                self.logger.normalInfo('设置进度为30')
                self.coverWindow.setValue(25)

                for row in range(0, rows_startForm):
                    for col in range(0, cols_startForm):
                        index_model = model_startForm.index(row, col)
                        cell = sheet_startForm.cell(row + 1, col + 1)
                        value = model_startForm.data(index_model, Qt.DisplayRole)
                        if value == 'None':
                            cell.value = ''
                        else:
                            cell.value = value

                self.logger.normalInfo('设置进度为40')
                self.coverWindow.setValue(35)

                self.logger.normalInfo('根据项目容器内容判别机柜类型，并合并单元格')
                value_add_1 = 35
                if project[4][0][1]['cabinet'] == '高密度':
                    for num_row in range(0, 10):
                        for num_column in range(0, 5):
                            value_add_1 += 0.4
                            sheet_startForm.merge_cells(start_row=4 + num_row * 12,
                                                        start_column=6 + num_column * 8,
                                                        end_row=4 + num_row * 12,
                                                        end_column=9 + num_column * 8)
                            sheet_startForm.merge_cells(start_row=5 + num_row * 12,
                                                        start_column=2 + num_column * 8,
                                                        end_row=5 + num_row * 12,
                                                        end_column=9 + num_column * 8)
                            sheet_startForm.merge_cells(start_row=6 + num_row * 12,
                                                        start_column=2 + num_column * 8,
                                                        end_row=6 + num_row * 12,
                                                        end_column=9 + num_column * 8)
                            self.coverWindow.setValue(value_add_1)

                elif project[4][0][1]['cabinet'] == '通用':
                    for num_row in range(0, 10):
                        for num_column in range(0, 5):
                            value_add_1 += 0.4
                            sheet_startForm.merge_cells(start_row=4 + num_row * 12,
                                                        start_column=5 + num_column * 6,
                                                        end_row=4 + num_row * 12,
                                                        end_column=7 + num_column * 6)
                            sheet_startForm.merge_cells(start_row=5 + num_row * 12,
                                                        start_column=2 + num_column * 6,
                                                        end_row=5 + num_row * 12,
                                                        end_column=7 + num_column * 6)
                            sheet_startForm.merge_cells(start_row=6 + num_row * 12,
                                                        start_column=2 + num_column * 6,
                                                        end_row=6 + num_row * 12,
                                                        end_column=7 + num_column * 6)
                            self.coverWindow.setValue(value_add_1)

                # 清册保存
                self.logger.normalInfo('保存清册页')
                num_listNum = len(project[5])
                value_add_2 = 0
                for index in range(0, num_listNum):
                    value_add_2 += 0.7
                    sheetName_IoForm = project[5][index][1]['name']
                    sheet_IoForm = wb.create_sheet(sheetName_IoForm)
                    model_IoForm = project[5][index][0]
                    rows_IoForm = model_IoForm.rowCount()
                    cols_IoForm = model_IoForm.columnCount()
                    for row in range(0, rows_IoForm):
                        for col in range(0, cols_IoForm):
                            index_model = model_IoForm.index(row, col)
                            cell = sheet_IoForm.cell(row + 1, col + 1)
                            value = model_IoForm.data(index_model, Qt.DisplayRole)
                            if value == 'None':
                                cell.value = ''
                            else:
                                cell.value = value
                    self.coverWindow.setValue(60+value_add_2/2)

                    self.logger.normalInfo('获取合并单元格容器内容合并单元格')
                    mergeList_IoForm = project[7][index][1]['mergeList']
                    self.logger.normalInfo('保存操作时合并清册单元格，单个单元格信息如下->')
                    try:
                        for mergeList in mergeList_IoForm:
                            if mergeList[2] == 0:
                                self.logger.abnormalInfo('当前行为块标题，不应合并|此时list内容为->\n' + str(mergeList))
                                continue
                            else:
                                self.logger.varInfo(str(mergeList))
                                sheet_IoForm.merge_cells(start_row=mergeList[0] + 1,
                                                         start_column=mergeList[1] + 1,
                                                         end_row=mergeList[0] + mergeList[2],
                                                         end_column=mergeList[1] + mergeList[3])
                    except:
                        self.logger.tranceInfo()
                    self.coverWindow.setValue(60 + value_add_2)

                wb.save(path)
                wb.close()

                self.logger.normalInfo('关闭掩盖窗体')
                self.coverWindow.setValue(100)
                self.coverWindow.timer.stop()
                self.coverWindow.close()

        self.logger.varInfo('FileSave')
        self.logger.varInfo('此时的容器内容->\n' + str(self.projectVessel.project))
        self.logger.varInfo('此时的总容器内容->\n' + str(self.allVessel))

        self.logger.normalInfo('保存完毕，牛逼')

    def FileSaveAsCel(self):
        self.logger.getFuncName('FileSaveAsCel')
        try:
            self.logger.normalInfo('尝试获取当前窗口名称')
            __, name = self.mdiObject.getCurrentWidgets()
        except Exception as e:
            self.logger.warringInfo(str(e))
            dlg = QMessageBox(self)
            dlg.setIcon(QMessageBox.Critical)
            dlg.setText(str(e))
            dlg.show()
            return 0
        else:
            if name is None:
                self.logger.warringInfo('当前无活动窗口')
                dlg = QMessageBox(self)
                dlg.setIcon(QMessageBox.Critical)
                dlg.setText('当前必须存在模板或清册窗口')
                dlg.show()
                return 0
            elif '清册' in name:
                projectName = name[4:]
            elif '模板' in name:
                projectName = name[3:]
            else:
                return 0
            self.logger.normalInfo('匹配项目名称')

        try:
            self.logger.normalInfo('弹出文件保存对话框')
            path, __ = QFileDialog.getSaveFileName(self, '另存为文件', '', '文件(*.xlsx)')
            self.logger.varInfo('对话框获取到的路径为->'+str(path))
            if str(path) == '':
                self.logger.abnormalInfo('对话框未获取到路径')
                return 0
        except Exception as e:
            self.logger.warringInfo(str(e))
            dlg = QMessageBox(self)
            dlg.setIcon(QMessageBox.Critical)
            dlg.setText(str(e))
            dlg.show()
            return 0
        else:
            self.logger.normalInfo('根据项目名称将路径保存进项目容器内')
            for project in self.allVessel:
                if projectName == project[0]:
                    project[8][0][0] = path

            self.logger.normalInfo('创建工作簿，并将其保存在相应路径下')
            wb = openpyxl.Workbook()
            wb.save(path)

            self.logger.normalInfo('进入文件保存')
            self.FileSave(projectName)

        self.logger.varInfo('FileSaveAsCel')
        self.logger.varInfo('此时的容器内容->\n' + str(self.projectVessel.project))
        self.logger.varInfo('此时的总容器内容->\n' + str(self.allVessel))

        self.logger.normalInfo('另存为操作完毕，叼哉')

    def EditAddCarton(self):
        self.logger.getFuncName('EditAddCarton')
        if self.allVessel == []:
            try:
                self.logger.normalInfo('创建自定义窗体管理对象，进入自定义卡件逻辑')
                self.childManager = PotsChildWindow_manager(self, self.dataVessel)
            except Exception as e:
                self.logger.tranceInfo()
                dlg = QMessageBox(self)
                dlg.setIcon(QMessageBox.Critical)
                dlg.setText(str(e))
                dlg.show()
        else:
            self.logger.abnormalInfo('有项目存在时不允许自定义卡件')
            dlg = QMessageBox(self)
            dlg.setIcon(QMessageBox.Critical)
            dlg.setText('有项目存在时不允许自定义卡件')
            dlg.show()
            return 0

    def IOSheetCreate(self):
        self.logger.getFuncName('IOSheetCreate')

        if self.allVessel == []:
            self.logger.abnormalInfo('此时尚未存在项目')
            dlg = QMessageBox(self)
            dlg.setIcon(QMessageBox.Critical)
            dlg.setText('无可生成项目')
            dlg.show()
            return 0

        try:
            self.logger.normalInfo('尝试获取当前窗口名称')
            __, name = self.mdiObject.getCurrentWidgets()
        except Exception as e:
            self.logger.warringInfo(str(e))
            dlg = QMessageBox(self)
            dlg.setIcon(QMessageBox.Critical)
            dlg.setText(str(e))
            dlg.show()
            return 0
        else:
            if name is None:
                self.logger.warringInfo('当前无活动窗口')
                dlg = QMessageBox(self)
                dlg.setIcon(QMessageBox.Critical)
                dlg.setText('当前必须存在模板或清册窗口')
                dlg.show()
                return 0
            elif '清册' in name:
                projectName = name[4:]
            elif '模板' in name:
                projectName = name[3:]
            else:
                return 0
            self.logger.normalInfo('匹配项目名称')

        self.logger.normalInfo('创建掩盖窗体')
        self.coverWindow = PotsCoverWindow(self)

        self.logger.normalInfo('设置进度为40')
        self.coverWindow.setValue(20)

        self.logger.normalInfo('根据各属性配置表单之间的联系，并以此更新清册内容')
        for project in self.allVessel:
            if project[0] == projectName:
                self.obj_confirm = PotsConfirmForms(project, self.dataVessel, self, self.treeObject, self.mdiObject,
                                                    self.coverWindow)
            else:
                continue

        self.logger.normalInfo('弹出清册生成完毕提示框')

        self.logger.normalInfo('设置进度为100')
        self.coverWindow.setValue(100)

        self.logger.normalInfo('关闭掩盖窗体')
        self.coverWindow.timer.stop()
        self.coverWindow.close()

        dlg = QMessageBox(self)
        dlg.setIcon(QMessageBox.Critical)
        dlg.setText('清册生成完毕')
        dlg.show()

        self.logger.varInfo('IOSheetCreate')
        self.logger.varInfo('此时的容器内容->\n' + str(self.projectVessel.project))
        self.logger.varInfo('此时的总容器内容->\n' + str(self.allVessel))

    def Analize(self):
        self.logger.getFuncName('Analize')
        dlg = QMessageBox(self)
        dlg.setIcon(QMessageBox.Critical)
        dlg.setText('这个还没做，不好意思@（*^*）!')
        dlg.show()

    def Check(self):
        self.logger.getFuncName('Check')
        dlg = QMessageBox(self)
        dlg.setIcon(QMessageBox.Critical)
        dlg.setText('这个还没做，不好意思@（*^*）!')
        dlg.show()

    def MachinCel(self):
        self.logger.getFuncName('MachinCel')
        dlg = QMessageBox(self)
        dlg.setIcon(QMessageBox.Critical)
        dlg.setText('这个还没做，不好意思@（*^*）!')
        dlg.show()


# 测试函数
def test():
    app = QApplication(sys.argv)
    UI = PotsCelUI()
    UI.show()
    if app.exec_():
        UI.logic.__db_currentCartons.close()
        UI.logic.__db_highDensityCartons.close()
        sys.exit()


# 开启动作
if __name__ == '__main__':
    loggerVar = PotsLogsVar()
    logger = PotsLogger('main')
    try:
        logger.normalInfo('创建应用')
        app = QApplication(sys.argv)
        logger.normalInfo('生成UI对象')
        UI = PotsCelUI()
        UI.showMaximized()
        logger.normalInfo('为应用添加样式')
        app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
        app.setStyleSheet(qdarkstyle.load_stylesheet(qt_api='pyqt5'))
        logger.normalInfo('展示UI')
        UI.show()
        if app.exec_():
            sys.exit()
    except:
        logger.tranceInfo()
