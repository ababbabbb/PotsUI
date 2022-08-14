# python包
from PyQt5.QtWidgets import QTableView, QApplication, QMenu
from PyQt5.QtGui import QStandardItem, QCursor
from PyQt5.QtCore import Qt
import openpyxl

# 自建包
from UIWidgets.formBasic import PotsFormCreatorBasic, PotsModel
from CodeTools.loggingTool import PotsLogger

# 类
class PotsTableView(QTableView):
    def __init__(self, parent):
        super().__init__(parent=parent)
        self.logger = PotsLogger('PotsTableView')
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        # 创建QMenu信号事件
        self.customContextMenuRequested.connect(self.showMenu)
        self.contextMenu = QMenu(self)
        self.CP = self.contextMenu.addAction('复制')
        self.JQ = self.contextMenu.addAction('剪切')
        self.NT = self.contextMenu.addAction('粘贴')
        self.DE = self.contextMenu.addAction('删除')
        self.CP.triggered.connect(self.copy)
        self.JQ.triggered.connect(self.cut)
        self.NT.triggered.connect(self.paste)
        self.DE.triggered.connect(self.delete)

    def delete(self):
        self.del_tb_text()

    def copy(self):
        text = self.selected_tb_text()
        if text:
            clipboard = QApplication.clipboard()
            clipboard.setText(text)

    def cut(self):
        self.copy()
        self.del_tb_text()

    def paste(self):
        self.paste_tb_text()

    def selected_tb_text(self):
        self.logger.getFuncName('selected_tb_text')
        try:
            self.logger.normalInfo('获取表格对象中被选中的数据索引列表')
            indexes = self.selectedIndexes()
            all_text = ''
            self.logger.normalInfo('遍历每个单元格')
            for index_math in indexes:
                row_math, col_math = index_math.row(), index_math.column()
                break

            for index_copy in indexes:
                self.logger.normalInfo(' 获取单元格的行号，列号')
                # row_toAdd, col_toAdd, value
                row, col = index_copy.row(), index_copy.column()
                value = self.model().data(index_copy, Qt.DisplayRole)
                if value is None or value == '':
                    value = 'xinxing14bi.'
                str_text = str(row-row_math)+'->'+str(col-col_math)+'->'+str(value)+'<-list,'
                all_text += str_text
            return all_text
        except:
            self.logger.tranceInfo()
            return ''

    def del_tb_text(self):
        self.logger.getFuncName('del_tb_text')
        try:
            indexes = self.selectedIndexes()
            for index in indexes:
                row, column = index.row(), index.column()
                model = self.model()
                item = QStandardItem()
                model.setItem(row, column, item)
            self.setModel(model)
        except:
            self.logger.tranceInfo()
            return 0

    def paste_tb_text(self):
        self.logger.getFuncName('paste_tb_text')
        try:
            indexes = self.selectedIndexes()
            for index in indexes:
                index = index
                break
            row_base, col_base = index.row(), index.column()

            text = QApplication.clipboard().text()
            if text == '':
                self.logger.abnormalInfo('数据不对，剪贴板里面的东西对不上->\n'+str(text))
                return 0

            list_text_0 = text.split("<-list,")
            list_text = []
            for list_cell_0 in list_text_0:
                if list_cell_0 == '' or list_cell_0 == ' ':
                    continue
                else:
                    list_cell = list_cell_0.split("->")
                    list_text.append(list_cell)

            for list_cell in list_text:
                if list_cell == ['']:
                    continue
                else:
                    if list_cell[2] == 'xinxing14bi.':
                        list_cell[2] = ''
                    value_cell = QStandardItem(list_cell[2])
                    self.model().setItem(row_base + int(list_cell[0]), col_base + int(list_cell[1]), value_cell)
        except:
            self.logger.tranceInfo()
            return 0

    def showMenu(self, pos):
        try:
            self.contextMenu.exec_(QCursor.pos())
        except:
            self.logger.tranceInfo()

    def keyPressEvent(self, event):
        """ Ctrl + C复制表格内容 """
        self.logger.getFuncName('keyPressEvent')
        try:
            if event.key() == Qt.Key_C and QApplication.keyboardModifiers() == Qt.ControlModifier:
                self.copy()
            elif event.key() == Qt.Key_X and QApplication.keyboardModifiers() == Qt.ControlModifier:
                self.cut()
            elif event.key() == Qt.Key_V and QApplication.keyboardModifiers() == Qt.ControlModifier:
                self.paste()
            elif event.key() == Qt.Key_Delete and QApplication.keyboardModifiers() == Qt.ControlModifier:
                self.delete()
            else:
                super().keyPressEvent(event)
        except:
            self.logger.tranceInfo()

class PotsIOFormBasic(PotsFormCreatorBasic):
    def __init__(self,
                 parent=None,
                 modelType=None,
                 modelName=None,
                 cabinet=None,
                 vessel=None,
                 isFiled=False):
        super().__init__(parent, modelType, modelName)
        self.logger = PotsLogger('PotsIOFormBasic')
        self.logger.normalInfo('IO表单创建对象初始化')

        self.counterregister = 0

        self.cabinet = cabinet
        self.vessel = vessel

        self.isFiled = isFiled

        self.checkSelf()

    def getDelegate(self, delegate):
        self.logger.getFuncName('getDelegate')
        self.logger.normalInfo('获取代理')

        self.delegate = delegate
        self.counterregister += 1

        self.logger.normalInfo('进入视图创建方法')
        self.createView()

    def checkSelf(self):
        self.logger.getFuncName('checkSelf')

        if self.parent is None:
            self.tipsDiag('无parent!')
        elif self.type_model is None:
            self.tipsDiag('无model类型!')
        elif self.name_model is None:
            self.tipsDiag('model无名称!')
        elif self.cabinet is None:
            self.tipsDiag('无机柜类型!')
        elif self.vessel is None:
            self.tipsDiag('无对象管理容器!')
        elif self.counterregister > 50:
            self.tipsDiag('本项目清册数量超限!')
        elif self.isFiled == True:
            self.logger.normalInfo('循环创建50个清册表单')
            while self.counterregister < 50:
                self.logger.normalInfo('创建表单->' + str(self.counterregister))
                self.counterregister += 1
                self.createModel()

            self.logger.normalInfo('将表达创建计数变量归零')
            self.counterregister = 0

            try:
                self.exitedData()
            except:
                self.logger.tranceInfo()

            self.logger.varInfo(str(self.vessel))
        else:
            self.logger.normalInfo('循环创建50个清册表单')
            while self.counterregister < 50:
                self.logger.normalInfo('创建表单->'+str(self.counterregister))
                self.counterregister += 1
                self.createModel()

            self.logger.normalInfo('将表达创建计数变量归零')
            self.counterregister = 0

            self.logger.varInfo(str(self.vessel))

    def exitedData(self):
        self.logger.getFuncName('exitedData')
        self.logger.normalInfo('通过遍历获取容器内保存的路径')
        for pathList in self.vessel.pathVessel:
            if self.name_model == pathList[1]['name']:
                path = pathList[0]
        self.logger.normalInfo('根据路径打开工作簿')
        wb = openpyxl.load_workbook(path)
        self.logger.normalInfo('遍历工作簿获取模板表单')
        for sheet in wb:
            if '清册' in sheet.title:
                IOModel_sheet = sheet
                index_IO = wb.index(sheet)
                IOModel = self.vessel.project[5][index_IO-1][0]
                rows = IOModel_sheet.max_row
                cols = IOModel_sheet.max_column
                self.logger.normalInfo('将表单内容按照坐标填入model')
                for r in range(0, rows):
                    for c in range(0, cols):
                        cell = IOModel_sheet.cell(r + 1, c + 1)
                        if str(cell.value) == 'None':
                            if c == 1:
                                n = 0
                                while(1):
                                    n += 1
                                    cell_carton = IOModel_sheet.cell(r + 1-n, c + 1)
                                    if str(cell_carton.value) == 'None':
                                        continue
                                    else:
                                        value = cell_carton.value
                                        break
                            else:
                                value = ''
                        else:
                            value = str(cell.value)
                        item = QStandardItem(value)
                        IOModel.setItem(r, c, item)
                merge = self.vessel.project[5][index_IO-1][1]['mergeList']
                mergeCells = sheet.merged_cells
                for merge_list in mergeCells:
                    if merge_list.min_row != merge_list.max_row and merge_list.min_col == merge_list.max_col:
                        list_merge = [merge_list.min_row-1,
                                      merge_list.min_col-1,
                                      merge_list.max_row-merge_list.min_row+1,
                                      merge_list.max_col-merge_list.min_col+1]
                        merge.append(list_merge)

            else:
                continue

        if IOModel_sheet is None:
            self.logger.abnormalInfo('工作簿内不存在清册表单')
            return 0

        self.logger.normalInfo('关闭已打开的工作簿')
        wb.close()

    def insertModelToVessel(self, model):
        self.logger.getFuncName('insertModelToVessel')

        """if self.counterregister <= 9:
                property_model = {
                'parent': self.parent,
                'type': 'model_' + self.type_model,
                'name': '清册0' + str(self.counterregister) + self.name_model,
                'cabinet': self.cabinet,
                'startFormSheetsIndex': self.counterregister,
                'mergeList': []
                }
            else:
                property_model = {
                'parent': self.parent,
                'type': 'model_' + self.type_model,
                'name': '清册' + str(self.counterregister) + self.name_model,
                'cabinet': self.cabinet,
                'startFormSheetsIndex': self.counterregister,
                'mergeList': []
                }
        self.vessel.IOmodelInsert(model, property_model)"""
        pass

    def insertViewToVessel(self, view):
        self.logger.getFuncName('insertViewToVessel')

        """if self.counterregister <= 9:
            property_view = {
                'parent': self.parent,
                'type': 'view_' + self.type_model,
                'name': '清册0' + str(self.counterregister) + self.name_model,
                'cabinet': self.cabinet,
                'startFormSheetsIndex': self.counterregister,
                'mergeList': [],
                'delegate': self.delegate
            }
        else:
            property_view = {
                'parent': self.parent,
                'type': 'view_' + self.type_model,
                'name': '清册' + str(self.counterregister) + self.name_model,
                'cabinet': self.cabinet,
                'startFormSheetsIndex': self.counterregister,
                'mergeList': [],
                'delegate': self.delegate
            }
        self.vessel.IOViewInsert(view, property_view)"""
        pass

    def createModel(self):
        self.logger.getFuncName('createModel')

        self.logger.normalInfo('创建model')
        model = PotsModel(2, 31, self.parent)
        IoRow1 = ['序号', '卡件', '线缆编号', '', '', '', '', '', 'DPU/DCS编号', '', '名称缩写', '设计编号', '测点名称',
                  '功能', 'I/O类型', '信号类型', '接点类型', '信号电源', '量程低限', '量程高限', '单位', '正常值',
                  '低三', '低二', '低一', '高一', '高二', '高三', '卷册号', '图号', '备注']
        IoRow2 = ['序号', '类型', '机柜', '槽位', '+', '-', 'C', '补', 'DCS编号', '数据连接位', '名称缩写', '设计编号',
                  'DCS编号', '项目测点名称', '功能', '种类', '信号形式', '供电来源', '低限', '高限', '单位', '报警',
                  '低三', '低二', '低一', '高一', '高二', '高三', '连接模块', '柜号', '备注']

        self.logger.normalInfo('向model内填入固定内容')
        for n in range(0, 31):
            R1C = QStandardItem(IoRow1[n])
            model.setItem(0, n, R1C)
            R2C = QStandardItem(IoRow2[n])
            model.setItem(1, n, R2C)

        # self.insertModelToVessel(model)
        self.logger.normalInfo('根据序列，编写model属性')
        property_model = {
            'parent': self.parent,
            'type': 'model_' + self.type_model,
            'name': '清册' + str(self.counterregister).zfill(2) + self.name_model,
            'cabinet': self.cabinet,
            'startFormSheetsIndex': self.counterregister,
            'mergeList': []
        }

        self.logger.normalInfo('将model及其属性填入容器')
        self.vessel.IOmodelInsert(model, property_model)

    def createView(self):
        self.logger.getFuncName('createView')
        self.logger.normalInfo('创建视图')
        view = PotsTableView(self.parent)
        view.setModel(self.vessel.modelVessel_IO[self.counterregister-1][0])
        view.setItemDelegate(self.delegate)

        # self.insertViewToVessel(view)

        self.logger.normalInfo('根据序列，编写view属性')
        property_view = {
            'parent': self.parent,
            'type': 'view_' + self.type_model,
            'name': '清册' + str(self.counterregister).zfill(2) + self.name_model,
            'cabinet': self.cabinet,
            'startFormSheetsIndex': self.counterregister,
            'mergeList': [],
            'delegate': self.delegate
        }

        self.logger.normalInfo('将view及其属性放入容器')
        self.vessel.IOViewInsert(view, property_view)

# 测试函数
def test():
    pass

# 开启动作
if __name__ == '__main__':
    test()
