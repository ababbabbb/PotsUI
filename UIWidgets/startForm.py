# python包
from PyQt5.QtWidgets import QTableView
from PyQt5.QtGui import QStandardItem
import openpyxl

# 自建包
from UIWidgets.formBasic import PotsFormCreatorBasic, PotsModel
from CodeTools.loggingTool import PotsLogger

#类
class PotsStartFormBasic(PotsFormCreatorBasic):
    def __init__(self,
                 parent=None,
                 modelType=None,
                 modelName=None,
                 cabinet=None,
                 vessel=None,
                 isFiled=False):
        super().__init__(parent, modelType, modelName)
        self.logger = PotsLogger('PotsStartFormBasic')
        self.logger.normalInfo('模件表单创建文件初始化')

        self.cabinet = cabinet
        self.vessel = vessel
        self.isFiled = isFiled

        self.checkSelf()

    def getDelegate(self, delegate):
        self.logger.getFuncName('getDelegate')
        self.logger.normalInfo('配置模件表单代理')
        self.delegate = delegate

        self.logger.normalInfo('进入视图创建方法')
        self.createView()

    def checkSelf(self):
        self.logger.getFuncName('checkSelf')
        self.logger.normalInfo('自检校验')

        if self.parent == None:
            self.tipsDiag('无parent!')
        elif self.type_model == None:
            self.tipsDiag('无model类型!')
        elif self.name_model == None:
            self.tipsDiag('model无名称!')
        elif self.cabinet == None:
            self.tipsDiag('无机柜类型!')
        elif self.vessel == None:
            self.tipsDiag('无对象管理容器!')
        elif self.isFiled == True:
            self.logger.normalInfo('打开模式下创建表单')
            self.logger.normalInfo('创建模型')
            self.createModel()
            self.logger.normalInfo('填入工作簿数据')
            self.exitedData()
        else:
            self.logger.normalInfo('新建模式下创建表单')
            self.logger.normalInfo('创建模型')
            self.createModel()

        self.logger.varInfo('checkSelf打印容器内容->\n'+str(self.vessel.project))

    def insertModelToVessel(self, model):
        self.logger.getFuncName('insertModelToVessel')
        self.logger.normalInfo('将model及其属性放入项目容器')
        property_model = {
            'parent': self.parent,
            'type': 'model_'+self.type_model,
            'name': '模板_'+self.name_model,
            'cabinet': self.cabinet
        }
        self.vessel.startModelInsert(model, property_model)

        self.logger.varInfo('model属性\n'+str(property_model))

    def insertViewToVessel(self, view):
        self.logger.getFuncName('insertViewToVessel')
        self.logger.normalInfo('将view及其属性放入项目容器')
        property_view = {
            'parent': self.parent,
            'type': 'view_' + self.type_model,
            'name': '模板_'+self.name_model,
            'cabinet': self.cabinet
        }
        self.vessel.startViewInsert(view, property_view)

        self.logger.varInfo('view属性\n'+str(property_view))

    def createModel(self):
        self.logger.getFuncName('createModel')
        self.logger.normalInfo('创建model')
        model = PotsModel(130, 60, self.parent)

        self.logger.normalInfo('填入固定内容')
        # 第一行固定内容
        self.insertValue(model, 0, 1, '起始表')
        self.insertValue(model, 0, 2, '结束表')
        self.insertValue(model, 0, 3, '名称缩写')
        self.insertValue(model, 0, 4, '设计编号')
        self.insertValue(model, 0, 5, '测点名称')
        self.insertValue(model, 0, 6, '信号类型')
        self.insertValue(model, 0, 7, '接点类型')
        self.insertValue(model, 0, 8, '信号电源')
        self.insertValue(model, 0, 9, '高量程')
        self.insertValue(model, 0, 10, '低量程')
        self.insertValue(model, 0, 11, '单位')
        self.insertValue(model, 0, 12, '高二值')
        self.insertValue(model, 0, 13, '高一值')
        self.insertValue(model, 0, 14, '低一值')
        self.insertValue(model, 0, 15, '低二值')
        self.insertValue(model, 0, 16, '结束列')
        self.insertValue(model, 0, 17, '常闭标识')
        self.insertValue(model, 0, 18, '供电标识')

        # 右侧固定内容

        # 左侧第一列固定内容
        for num in range(0, 10):
            self.insertValue(model, 4 + (num * 12), 0, 'DPU')
            self.insertValue(model, 5 + (num * 12), 0, 'CAB')
            self.insertValue(model, 6 + (num * 12), 0, 'BRC')
            self.insertValue(model, 7 + (num * 12), 0, 'SL1')
            self.insertValue(model, 8 + (num * 12), 0, 'SL2')
            self.insertValue(model, 9 + (num * 12), 0, 'SL3')
            self.insertValue(model, 10 + (num * 12), 0, 'SL4')
            self.insertValue(model, 11 + (num * 12), 0, 'SL5')
            self.insertValue(model, 12 + (num * 12), 0, 'SL6')
            self.insertValue(model, 13 + (num * 12), 0, 'SL7')
            self.insertValue(model, 14 + (num * 12), 0, 'SL8')

        if self.cabinet == '高密度':
            num_sheet = 0
            for num_row in range(0, 10):
                for num_column in range(0, 5):
                    num_sheet += 1
                    self.insertValue(model, 3+num_row*12, 1+num_column*8, '第')
                    self.insertValue(model, 3+num_row*12, 2+num_column*8, str(num_sheet))
                    self.insertValue(model, 3+num_row*12, 3+num_column*8, '表')
                    self.insertValue(model, 3+num_row*12, 4+num_column*8, '(H)')
        elif self.cabinet == '通用':
            num_sheet = 0
            for num_row in range(0, 10):
                for num_column in range(0, 5):
                    num_sheet += 1
                    self.insertValue(model, 3 + num_row * 12, 1 + num_column * 6, '第')
                    self.insertValue(model, 3 + num_row * 12, 2 + num_column * 6, str(num_sheet))
                    self.insertValue(model, 3 + num_row * 12, 3 + num_column * 6, '表')
        else:
            self.logger.abnormalInfo('机柜类型未知')
            self.tipsDiag('机柜类型陷入了未知情况')

        self.logger.normalInfo('将model放入容器')
        self.insertModelToVessel(model)

        self.logger.normalInfo('将model转为全局变量')
        self.model = model

    def exitedData(self):
        self.logger.getFuncName('exitedData')
        self.logger.normalInfo('通过遍历获取容器内保存的路径')
        for pathList in self.vessel.pathVessel:
            if self.name_model == pathList[1]['name']:
                path = pathList[0]
        self.logger.normalInfo('根据路径打开工作簿')
        wb = openpyxl.load_workbook(path)
        self.logger.normalInfo('遍历工作簿获取模板表单')
        for sheet in wb:
            if '模件' in sheet.title:
                startModel_sheet = sheet
                break
            else:
                continue

        if startModel_sheet is None:
            self.logger.abnormalInfo('工作簿内不存在模件表单')
            return 0

        rows = startModel_sheet.max_row
        cols = startModel_sheet.max_column
        self.logger.normalInfo('将表单内容按照坐标填入model')
        for r in range(0, rows):
            for c in range(0, cols):
                cell = startModel_sheet.cell(r + 1, c + 1)
                if str(cell.value) == 'None':
                    value = ''
                else:
                    value = str(cell.value)
                item = QStandardItem(value)
                self.model.setItem(r, c, item)

        self.logger.normalInfo('关闭已打开的工作簿')
        wb.close()

    def createView(self):
        self.logger.getFuncName('createView')
        self.logger.normalInfo('创建视图')
        view = QTableView(self.parent)
        view.setModel(self.model)

        self.logger.normalInfo('根据机柜类型设置视图代理，并合并单元格')
        if self.cabinet == '高密度':
            view.setItemDelegate(self.delegate)
            for num_row in range(0, 10):
                for num_column in range(0, 5):
                    view.setSpan(3 + num_row * 12, 5 + num_column * 8, 1, 4)
                    view.setSpan(4 + num_row * 12, 1 + num_column * 8, 1, 8)
                    view.setSpan(5 + num_row * 12, 1 + num_column * 8, 1, 8)
        elif self.cabinet == '通用':
            view.setItemDelegate(self.delegate)
            for num_row in range(0, 10):
                for num_column in range(0, 5):
                    view.setSpan(3 + num_row * 12, 4 + num_column * 6, 1, 3)
                    view.setSpan(4 + num_row * 12, 1 + num_column * 6, 1, 6)
                    view.setSpan(5 + num_row * 12, 1 + num_column * 6, 1, 6)
        else:
            pass

        self.logger.normalInfo('将视图放入容器')
        self.insertViewToVessel(view)

        self.logger.normalInfo('将view专为全局变量')
        self.view = view

# 测试函数
def test():
    pass

#开启动作
if __name__ == '__main__':
    test()
