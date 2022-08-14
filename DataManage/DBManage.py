# 说明
"""
用于数据库（也许我该用基本表格而非db）管理
"""

# python包
import os
import openpyxl

# 自建包
from CodeTools.loggingTool import PotsLogger

# 类
class PotsDataVessel():
    def __init__(self):
        self.logger = PotsLogger('PotsBaseManageObj')

        path_exe = os.getcwd()
        self.baseFilePath = path_exe+'/configData'
        if os.path.exists(self.baseFilePath):
            pass
        else:
            os.mkdir("configData")

        self.basePath = self.baseFilePath+'/cartons.xlsx'
        if os.path.exists(self.basePath):
            self.wb_carton = openpyxl.load_workbook(self.basePath)
            self.wb_carton.save(self.basePath)
            self.wb_carton.close()
        else:
            self.wb_carton = openpyxl.Workbook()
            del self.wb_carton['Sheet']
            self.sheet_fixedCarton = self.wb_carton.create_sheet('卡件汇总(固定)', 0)   # 卡件、支持高密度、通道数、线制
            self.sheet_customizeCarton = self.wb_carton.create_sheet('卡件汇总(自定义)', 1)   # 卡件、支持高密度、通道数、线制
            self.sheet_KM231A = self.wb_carton.create_sheet('KM231A', 2)
            self.sheet_KM231B = self.wb_carton.create_sheet('KM231B', 3)
            self.sheet_KM231C = self.wb_carton.create_sheet('KM231C', 4)
            self.sheet_KM231D = self.wb_carton.create_sheet('KM231D', 5)
            self.sheet_KM231E = self.wb_carton.create_sheet('KM231E', 6)
            self.sheet_KM231F = self.wb_carton.create_sheet('KM231F', 7)
            self.sheet_KM232A = self.wb_carton.create_sheet('KM232A', 8)
            self.sheet_KM233A = self.wb_carton.create_sheet('KM233A', 9)
            self.sheet_KM233B = self.wb_carton.create_sheet('KM233B', 10)
            self.sheet_KM234A = self.wb_carton.create_sheet('KM234A', 11)
            self.sheet_KM235A = self.wb_carton.create_sheet('KM235A', 12)
            self.sheet_KM235B = self.wb_carton.create_sheet('KM235B', 13)
            self.sheet_KM236A = self.wb_carton.create_sheet('KM236A', 14)
            self.sheet_KM236B = self.wb_carton.create_sheet('KM236B', 15)
            self.sheet_KM237A = self.wb_carton.create_sheet('KM237A', 16)
            self.sheet_KM331A = self.wb_carton.create_sheet('KM331A', 17)
            self.sheet_KM331B = self.wb_carton.create_sheet('KM331B', 18)
            self.sheet_KM331C = self.wb_carton.create_sheet('KM331C', 19)
            self.sheet_KM331E = self.wb_carton.create_sheet('KM331E', 20)
            self.sheet_KM332A = self.wb_carton.create_sheet('KM332A', 21)
            self.sheet_KM333A = self.wb_carton.create_sheet('KM333A', 22)
            self.sheet_KM333B = self.wb_carton.create_sheet('KM333B', 23)
            self.sheet_KM334A = self.wb_carton.create_sheet('KM334A', 24)
            self.sheet_KM335A = self.wb_carton.create_sheet('KM335A', 25)
            self.sheet_KM335B = self.wb_carton.create_sheet('KM335B', 26)
            self.sheet_KM336A = self.wb_carton.create_sheet('KM336A', 27)
            self.sheet_KM531A = self.wb_carton.create_sheet('KM531A', 28)
            self.sheet_KM532A = self.wb_carton.create_sheet('KM532A', 29)
            self.sheet_KM532B = self.wb_carton.create_sheet('KM532B', 30)
            self.sheet_KM533A = self.wb_carton.create_sheet('KM533A', 31)
            self.sheet_KM534A = self.wb_carton.create_sheet('KM534A', 32)
            self.sheet_KM535A = self.wb_carton.create_sheet('KM535A', 33)
            self.sheet_KM536A = self.wb_carton.create_sheet('KM536A', 34)
            self.sheet_KM631A = self.wb_carton.create_sheet('KM631A', 35)
            self.sheet_KM631B = self.wb_carton.create_sheet('KM631B', 36)
            self.sheet_KM631C = self.wb_carton.create_sheet('KM631C', 37)
            self.sheet_KM632A = self.wb_carton.create_sheet('KM632A', 38)
            self.sheet_KM632C = self.wb_carton.create_sheet('KM632C', 39)
            self.sheet_KM633A = self.wb_carton.create_sheet('KM633A', 40)
            self.sheet_KM633B = self.wb_carton.create_sheet('KM633B', 41)
            self.sheet_KB431A = self.wb_carton.create_sheet('KB431A', 42)
            self.sheet_KB432A = self.wb_carton.create_sheet('KB432A', 43)
            self.sheet_KB432C = self.wb_carton.create_sheet('KB432C', 44)
            self.sheet_KB432D = self.wb_carton.create_sheet('KB432D', 45)

            self.insertValueToCartonSheet(self.sheet_fixedCarton, 1, 'KM231A', 'AI', 'True', '8', ['2线制', '4线制'])
            self.insertValueToSheet(self.sheet_KM231A, '2线制', 1, ['1', '2'])
            self.insertValueToSheet(self.sheet_KM231A, '2线制', 2, ['23', '24'])
            self.insertValueToSheet(self.sheet_KM231A, '2线制', 3, ['6', '7'])
            self.insertValueToSheet(self.sheet_KM231A, '2线制', 4, ['28', '29'])
            self.insertValueToSheet(self.sheet_KM231A, '2线制', 5, ['11', '12'])
            self.insertValueToSheet(self.sheet_KM231A, '2线制', 6, ['33', '34'])
            self.insertValueToSheet(self.sheet_KM231A, '2线制', 7, ['16', '17'])
            self.insertValueToSheet(self.sheet_KM231A, '2线制', 8, ['38', '39'])
            self.insertValueToSheet(self.sheet_KM231A, '4线制', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM231A, '4线制', 2, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM231A, '4线制', 3, ['7', '8'])
            self.insertValueToSheet(self.sheet_KM231A, '4线制', 4, ['29', '30'])
            self.insertValueToSheet(self.sheet_KM231A, '4线制', 5, ['12', '13'])
            self.insertValueToSheet(self.sheet_KM231A, '4线制', 6, ['34', '35'])
            self.insertValueToSheet(self.sheet_KM231A, '4线制', 7, ['17', '18'])
            self.insertValueToSheet(self.sheet_KM231A, '4线制', 8, ['39', '40'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 2, 'KM231B', 'AI', 'True', '16', ['2线制', '4线制', '220V'])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 1, ['1', '2'])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 2, ['21', '22'])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 3, [])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 4, [])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 5, [])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 6, [])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 7, [])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 8, [])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 9, ['11', '12'])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 10, ['31', '32'])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 11, [])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 12, [])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 13, [])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 14, [])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 15, [])
            self.insertValueToSheet(self.sheet_KM231B, '2线制', 16, [])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 2, ['22', '23'])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 3, ['4', '5'])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 4, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 5, ['6', '7'])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 6, ['26', '27'])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 7, ['8', '9'])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 8, ['28', '29'])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 9, ['12', '13'])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 10, ['32', '33'])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 11, ['14', '15'])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 12, ['34', '35'])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 13, ['16', '17'])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 14, ['36', '37'])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 15, ['18', '19'])
            self.insertValueToSheet(self.sheet_KM231B, '4线制', 16, ['38', '39'])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 1, [])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 2, [])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 3, [])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 4, [])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 5, [])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 6, [])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 7, [])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 8, ['30', '29'])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 9, [])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 10, [])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 11, [])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 12, [])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 13, [])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 14, [])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 15, [])
            self.insertValueToSheet(self.sheet_KM231B, '220V', 16, ['40', '39'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 3, 'KM231C', 'AI', 'True', '16', ['4线制', '220V'])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 2, ['4', '5'])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 3, ['7', '8'])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 4, ['9', '10'])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 5, ['12', '13'])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 6, ['14', '15'])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 7, ['17', '18'])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 8, ['19', '20'])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 9, [])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 10, [])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 11, [])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 12, [])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 13, [])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 14, [])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 15, [])
            self.insertValueToSheet(self.sheet_KM231C, '4线制', 16, [])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 1, [])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 2, [])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 3, [])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 4, [])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 5, [])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 6, [])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 7, [])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 8, [])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 9, ['22', '23'])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 10, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 11, ['27', '28'])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 12, ['29', '30'])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 13, ['32', '33'])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 14, ['34', '35'])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 15, ['37', '38'])
            self.insertValueToSheet(self.sheet_KM231C, '220V', 16, ['39', '40'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 4, 'KM231D', 'AI', 'True', '8', ['2线制', '4线制', '220V'])
            self.insertValueToSheet(self.sheet_KM231D, '2线制', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM231D, '2线制', 2, ['22', '23'])
            self.insertValueToSheet(self.sheet_KM231D, '2线制', 3, ['7', '8'])
            self.insertValueToSheet(self.sheet_KM231D, '2线制', 4, ['27', '28'])
            self.insertValueToSheet(self.sheet_KM231D, '2线制', 5, ['12', '13'])
            self.insertValueToSheet(self.sheet_KM231D, '2线制', 6, ['32', '33'])
            self.insertValueToSheet(self.sheet_KM231D, '2线制', 7, ['17', '18'])
            self.insertValueToSheet(self.sheet_KM231D, '2线制', 8, ['37', '38'])
            self.insertValueToSheet(self.sheet_KM231D, '4线制', 1, ['3', '4'])
            self.insertValueToSheet(self.sheet_KM231D, '4线制', 2, ['23', '24'])
            self.insertValueToSheet(self.sheet_KM231D, '4线制', 3, ['8', '9'])
            self.insertValueToSheet(self.sheet_KM231D, '4线制', 4, ['28', '29'])
            self.insertValueToSheet(self.sheet_KM231D, '4线制', 5, ['13', '14'])
            self.insertValueToSheet(self.sheet_KM231D, '4线制', 6, ['33', '34'])
            self.insertValueToSheet(self.sheet_KM231D, '4线制', 7, ['18', '19'])
            self.insertValueToSheet(self.sheet_KM231D, '4线制', 8, ['38', '39'])
            self.insertValueToSheet(self.sheet_KM231D, '220V', 1, ['5', '4'])
            self.insertValueToSheet(self.sheet_KM231D, '220V', 2, ['25', '24'])
            self.insertValueToSheet(self.sheet_KM231D, '220V', 3, ['10', '9'])
            self.insertValueToSheet(self.sheet_KM231D, '220V', 4, ['30', '29'])
            self.insertValueToSheet(self.sheet_KM231D, '220V', 5, ['15', '14'])
            self.insertValueToSheet(self.sheet_KM231D, '220V', 6, ['35', '34'])
            self.insertValueToSheet(self.sheet_KM231D, '220V', 7, ['20', '19'])
            self.insertValueToSheet(self.sheet_KM231D, '220V', 8, ['40', '39'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 5, 'KM231E', 'AI', 'True', '8', ['2线制', '4线制', '220V'])
            self.insertValueToSheet(self.sheet_KM231E, '2线制', 1, ['21', '22'])
            self.insertValueToSheet(self.sheet_KM231E, '2线制', 2, ['4', '5'])
            self.insertValueToSheet(self.sheet_KM231E, '2线制', 3, ['26', '28'])
            self.insertValueToSheet(self.sheet_KM231E, '2线制', 4, ['9', '11'])
            self.insertValueToSheet(self.sheet_KM231E, '2线制', 5, ['31', '33'])
            self.insertValueToSheet(self.sheet_KM231E, '2线制', 6, ['14', '16'])
            self.insertValueToSheet(self.sheet_KM231E, '2线制', 7, ['36', '38'])
            self.insertValueToSheet(self.sheet_KM231E, '2线制', 8, ['18', '20'])
            self.insertValueToSheet(self.sheet_KM231E, '4线制', 1, ['22', '23'])
            self.insertValueToSheet(self.sheet_KM231E, '4线制', 2, ['5', '6'])
            self.insertValueToSheet(self.sheet_KM231E, '4线制', 3, ['27', '28'])
            self.insertValueToSheet(self.sheet_KM231E, '4线制', 4, ['10', '11'])
            self.insertValueToSheet(self.sheet_KM231E, '4线制', 5, ['32', '33'])
            self.insertValueToSheet(self.sheet_KM231E, '4线制', 6, ['15', '16'])
            self.insertValueToSheet(self.sheet_KM231E, '4线制', 7, ['37', '38'])
            self.insertValueToSheet(self.sheet_KM231E, '4线制', 8, ['19', '20'])
            self.insertValueToSheet(self.sheet_KM231E, '220V', 1, ['22', '23'])
            self.insertValueToSheet(self.sheet_KM231E, '220V', 2, ['5', '6'])
            self.insertValueToSheet(self.sheet_KM231E, '220V', 3, ['27', '28'])
            self.insertValueToSheet(self.sheet_KM231E, '220V', 4, ['10', '11'])
            self.insertValueToSheet(self.sheet_KM231E, '220V', 5, ['32', '33'])
            self.insertValueToSheet(self.sheet_KM231E, '220V', 6, ['15', '16'])
            self.insertValueToSheet(self.sheet_KM231E, '220V', 7, ['37', '38'])
            self.insertValueToSheet(self.sheet_KM231E, '220V', 8, ['19', '20'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 6, 'KM231F', 'AI', 'True', '8', ['2线制', '4线制'])
            self.insertValueToSheet(self.sheet_KM231F, '2线制', 1, ['1', '2'])
            self.insertValueToSheet(self.sheet_KM231F, '2线制', 2, ['23', '24'])
            self.insertValueToSheet(self.sheet_KM231F, '2线制', 3, ['6', '7',])
            self.insertValueToSheet(self.sheet_KM231F, '2线制', 4, ['28', '29'])
            self.insertValueToSheet(self.sheet_KM231F, '2线制', 5, ['11', '12'])
            self.insertValueToSheet(self.sheet_KM231F, '2线制', 6, ['33', '34'])
            self.insertValueToSheet(self.sheet_KM231F, '2线制', 7, ['16', '17'])
            self.insertValueToSheet(self.sheet_KM231F, '2线制', 8, ['38', '39'])
            self.insertValueToSheet(self.sheet_KM231F, '4线制', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM231F, '4线制', 2, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM231F, '4线制', 3, ['7', '8'])
            self.insertValueToSheet(self.sheet_KM231F, '4线制', 4, ['29', '30'])
            self.insertValueToSheet(self.sheet_KM231F, '4线制', 5, ['12', '13'])
            self.insertValueToSheet(self.sheet_KM231F, '4线制', 6, ['34', '35'])
            self.insertValueToSheet(self.sheet_KM231F, '4线制', 7, ['17', '18'])
            self.insertValueToSheet(self.sheet_KM231F, '4线制', 8, ['39', '40'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 7, 'KM232A', 'RT', 'True', '8', ['2线制', '4线制'])
            self.insertValueToSheet(self.sheet_KM232A, '2线制', 1, ['3', '4', '5'])
            self.insertValueToSheet(self.sheet_KM232A, '2线制', 2, ['23', '24', '25'])
            self.insertValueToSheet(self.sheet_KM232A, '2线制', 3, ['8', '9', '10'])
            self.insertValueToSheet(self.sheet_KM232A, '2线制', 4, ['28', '29', '30'])
            self.insertValueToSheet(self.sheet_KM232A, '2线制', 5, ['13', '14', '15'])
            self.insertValueToSheet(self.sheet_KM232A, '2线制', 6, ['33', '34', '35'])
            self.insertValueToSheet(self.sheet_KM232A, '2线制', 7, ['18', '19', '20'])
            self.insertValueToSheet(self.sheet_KM232A, '2线制', 8, ['38', '39', '40'])
            self.insertValueToSheet(self.sheet_KM232A, '4线制', 1, ['3', '4', '5'])
            self.insertValueToSheet(self.sheet_KM232A, '4线制', 2, ['23', '24', '25'])
            self.insertValueToSheet(self.sheet_KM232A, '4线制', 3, ['8', '9', '10'])
            self.insertValueToSheet(self.sheet_KM232A, '4线制', 4, ['28', '29', '30'])
            self.insertValueToSheet(self.sheet_KM232A, '4线制', 5, ['13', '14', '15'])
            self.insertValueToSheet(self.sheet_KM232A, '4线制', 6, ['33', '34', '35'])
            self.insertValueToSheet(self.sheet_KM232A, '4线制', 7, ['18', '19', '20'])
            self.insertValueToSheet(self.sheet_KM232A, '4线制', 8, ['38', '39', '40'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 8, 'KM233A', 'TC', 'True', '8', ['2线制'])
            self.insertValueToSheet(self.sheet_KM233A, '2线制', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM233A, '2线制', 2, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM233A, '2线制', 3, ['6', '7'])
            self.insertValueToSheet(self.sheet_KM233A, '2线制', 4, ['28', '29'])
            self.insertValueToSheet(self.sheet_KM233A, '2线制', 5, ['10', '11'])
            self.insertValueToSheet(self.sheet_KM233A, '2线制', 6, ['32', '33'])
            self.insertValueToSheet(self.sheet_KM233A, '2线制', 7, ['14', '15'])
            self.insertValueToSheet(self.sheet_KM233A, '2线制', 8, ['36', '37'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 9, 'KM233B', 'TC', 'True', '16', ['2线制'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 1, ['1', '2'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 2, ['21', '22'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 3, ['3', '4'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 4, ['23', '24'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 5, ['5', '6'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 6, ['25', '26'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 7, ['14', '15'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 8, ['36', '37'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 9, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 10, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 11, ['6', '7'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 12, ['28', '29'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 13, ['10', '11'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 14, ['32', '33'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 15, ['14', '15'])
            self.insertValueToSheet(self.sheet_KM233B, '2线制', 16, ['36', '37'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 10, 'KM234A', 'DI', 'True', '16', ['2线制'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 1, ['2', '22'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 2, ['3', '23'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 3, ['4', '24'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 4, ['5', '25'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 5, ['7', '27'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 6, ['8', '28'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 7, ['9', '29'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 8, ['10', '30'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 9, ['12', '32'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 10, ['13', '33'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 11, ['14', '34'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 12, ['15', '35'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 13, ['17', '37'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 14, ['18', '38'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 15, ['19', '39'])
            self.insertValueToSheet(self.sheet_KM234A, '2线制', 16, ['20', '40'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 11, 'KM235A', 'DO', 'True', '8', ['2线制'])
            self.insertValueToSheet(self.sheet_KM235A, '2线制', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM235A, '2线制', 2, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM235A, '2线制', 3, ['7', '8'])
            self.insertValueToSheet(self.sheet_KM235A, '2线制', 4, ['29', '30'])
            self.insertValueToSheet(self.sheet_KM235A, '2线制', 5, ['12', '13'])
            self.insertValueToSheet(self.sheet_KM235A, '2线制', 6, ['34', '35'])
            self.insertValueToSheet(self.sheet_KM235A, '2线制', 7, ['17', '18'])
            self.insertValueToSheet(self.sheet_KM235A, '2线制', 8, ['39', '40'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 12, 'KM235B', 'DO', 'True', '16', ['2线制'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 1, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 2, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 3, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 4, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 5, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 6, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 7, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 8, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 9, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 10, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 11, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 12, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 13, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 14, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 15, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM235B, '2线制', 16, ['转接', '至', '组件'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 13, 'KM236A', 'AO', 'True', '6', ['2线制'])
            self.insertValueToSheet(self.sheet_KM236A, '2线制', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM236A, '2线制', 2, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM236A, '2线制', 3, ['7', '8'])
            self.insertValueToSheet(self.sheet_KM236A, '2线制', 4, ['29', '30'])
            self.insertValueToSheet(self.sheet_KM236A, '2线制', 5, ['12', '13'])
            self.insertValueToSheet(self.sheet_KM236A, '2线制', 6, ['34', '35'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 14, 'KM236B', 'AO', 'True', '4', ['2线制', '4线制'])
            self.insertValueToSheet(self.sheet_KM236B, '2线制', 1, ['4', '2'])
            self.insertValueToSheet(self.sheet_KM236B, '2线制', 2, ['26', '24'])
            self.insertValueToSheet(self.sheet_KM236B, '2线制', 3, ['9', '7'])
            self.insertValueToSheet(self.sheet_KM236B, '2线制', 4, ['31', '29'])
            self.insertValueToSheet(self.sheet_KM236B, '4线制', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM236B, '4线制', 2, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM236B, '4线制', 3, ['7', '8'])
            self.insertValueToSheet(self.sheet_KM236B, '4线制', 4, ['29', '30'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 15, 'KM237A', 'HZ', 'True', '8', ['OC', 'TTL'])
            self.insertValueToSheet(self.sheet_KM237A, 'OC', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM237A, 'OC', 2, ['7', '8'])
            self.insertValueToSheet(self.sheet_KM237A, 'OC', 3, ['12', '13'])
            self.insertValueToSheet(self.sheet_KM237A, 'OC', 4, ['17', '18'])
            self.insertValueToSheet(self.sheet_KM237A, 'OC', 5, [])
            self.insertValueToSheet(self.sheet_KM237A, 'OC', 6, [])
            self.insertValueToSheet(self.sheet_KM237A, 'OC', 7, [])
            self.insertValueToSheet(self.sheet_KM237A, 'OC', 8, [])
            self.insertValueToSheet(self.sheet_KM237A, 'TTL', 1, [])
            self.insertValueToSheet(self.sheet_KM237A, 'TTL', 2, [])
            self.insertValueToSheet(self.sheet_KM237A, 'TTL', 3, [])
            self.insertValueToSheet(self.sheet_KM237A, 'TTL', 4, [])
            self.insertValueToSheet(self.sheet_KM237A, 'TTL', 5, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM237A, 'TTL', 6, ['29', '30'])
            self.insertValueToSheet(self.sheet_KM237A, 'TTL', 7, ['34', '35'])
            self.insertValueToSheet(self.sheet_KM237A, 'TTL', 8, ['39', '40'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 16, 'KM331A', 'AI', 'False', '8', ['2线制', '4线制'])
            self.insertValueToSheet(self.sheet_KM331A, '2线制', 1, ['1', '2'])
            self.insertValueToSheet(self.sheet_KM331A, '2线制', 2, ['23', '24'])
            self.insertValueToSheet(self.sheet_KM331A, '2线制', 3, ['6', '7'])
            self.insertValueToSheet(self.sheet_KM331A, '2线制', 4, ['28', '29'])
            self.insertValueToSheet(self.sheet_KM331A, '2线制', 5, ['11', '12'])
            self.insertValueToSheet(self.sheet_KM331A, '2线制', 6, ['33', '34'])
            self.insertValueToSheet(self.sheet_KM331A, '2线制', 7, ['16', '17'])
            self.insertValueToSheet(self.sheet_KM331A, '2线制', 8, ['38', '39'])
            self.insertValueToSheet(self.sheet_KM331A, '4线制', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM331A, '4线制', 2, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM331A, '4线制', 3, ['7', '8'])
            self.insertValueToSheet(self.sheet_KM331A, '4线制', 4, ['29', '30'])
            self.insertValueToSheet(self.sheet_KM331A, '4线制', 5, ['12', '13'])
            self.insertValueToSheet(self.sheet_KM331A, '4线制', 6, ['34', '35'])
            self.insertValueToSheet(self.sheet_KM331A, '4线制', 7, ['17', '18'])
            self.insertValueToSheet(self.sheet_KM331A, '4线制', 8, ['39', '40'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 17, 'KM331B', 'AI', 'False', '16', ['2线制', '4线制', '220V'])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 1, ['1', '2'])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 2, ['21', '22'])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 3, [])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 4, [])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 5, [])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 6, [])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 7, [])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 8, [])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 9, ['11', '12'])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 10, ['31', '32'])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 11, [])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 12, [])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 13, [])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 14, [])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 15, [])
            self.insertValueToSheet(self.sheet_KM331B, '2线制', 16, [])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 2, ['22', '23'])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 3, ['4', '5'])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 4, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 5, ['6', '7'])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 6, ['26', '27'])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 7, ['8', '9'])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 8, ['28', '29'])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 9, ['12', '13'])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 10, ['32', '33'])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 11, ['14', '15'])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 12, ['34', '35'])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 13, ['16', '17'])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 14, ['36', '37'])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 15, ['18', '19'])
            self.insertValueToSheet(self.sheet_KM331B, '4线制', 16, ['38', '39'])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 1, [])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 2, [])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 3, [])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 4, [])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 5, [])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 6, [])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 7, [])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 8, ['30', '29'])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 9, [])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 10, [])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 11, [])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 12, [])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 13, [])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 14, [])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 15, [])
            self.insertValueToSheet(self.sheet_KM331B, '220V', 16, ['40', '39'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 18, 'KM331C', 'AI', 'False', '16', ['4线制', '220V'])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 2, ['4', '5'])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 3, ['7', '8'])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 4, ['9', '10'])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 5, ['12', '13'])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 6, ['14', '15'])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 7, ['17', '18'])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 8, ['19', '20'])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 9, [])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 10, [])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 11, [])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 12, [])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 13, [])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 14, [])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 15, [])
            self.insertValueToSheet(self.sheet_KM331C, '4线制', 16, [])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 1, [])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 2, [])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 3, [])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 4, [])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 5, [])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 6, [])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 7, [])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 8, [])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 9, ['22', '23'])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 10, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 11, ['27', '28'])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 12, ['29', '30'])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 13, ['32', '33'])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 14, ['34', '35'])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 15, ['37', '38'])
            self.insertValueToSheet(self.sheet_KM331C, '220V', 16, ['39', '40'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 19, 'KM331E', 'AI', 'False', '8', ['2线制', '4线制', '220V'])
            self.insertValueToSheet(self.sheet_KM331E, '2线制', 1, ['21', '22'])
            self.insertValueToSheet(self.sheet_KM331E, '2线制', 2, ['4', '5'])
            self.insertValueToSheet(self.sheet_KM331E, '2线制', 3, ['26', '28'])
            self.insertValueToSheet(self.sheet_KM331E, '2线制', 4, ['9', '11'])
            self.insertValueToSheet(self.sheet_KM331E, '2线制', 5, ['31', '33'])
            self.insertValueToSheet(self.sheet_KM331E, '2线制', 6, ['14', '16'])
            self.insertValueToSheet(self.sheet_KM331E, '2线制', 7, ['36', '38'])
            self.insertValueToSheet(self.sheet_KM331E, '2线制', 8, ['18', '20'])
            self.insertValueToSheet(self.sheet_KM331E, '4线制', 1, ['22', '23'])
            self.insertValueToSheet(self.sheet_KM331E, '4线制', 2, ['5', '6'])
            self.insertValueToSheet(self.sheet_KM331E, '4线制', 3, ['27', '28'])
            self.insertValueToSheet(self.sheet_KM331E, '4线制', 4, ['10', '11'])
            self.insertValueToSheet(self.sheet_KM331E, '4线制', 5, ['32', '33'])
            self.insertValueToSheet(self.sheet_KM331E, '4线制', 6, ['15', '16'])
            self.insertValueToSheet(self.sheet_KM331E, '4线制', 7, ['37', '38'])
            self.insertValueToSheet(self.sheet_KM331E, '4线制', 8, ['19', '20'])
            self.insertValueToSheet(self.sheet_KM331E, '220V', 1, ['22', '23'])
            self.insertValueToSheet(self.sheet_KM331E, '220V', 2, ['5', '6'])
            self.insertValueToSheet(self.sheet_KM331E, '220V', 3, ['27', '28'])
            self.insertValueToSheet(self.sheet_KM331E, '220V', 4, ['10', '11'])
            self.insertValueToSheet(self.sheet_KM331E, '220V', 5, ['32', '33'])
            self.insertValueToSheet(self.sheet_KM331E, '220V', 6, ['15', '16'])
            self.insertValueToSheet(self.sheet_KM331E, '220V', 7, ['37', '38'])
            self.insertValueToSheet(self.sheet_KM331E, '220V', 8, ['19', '20'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 20, 'KM332A', 'RT', 'False', '8', ['2线制', '4线制'])
            self.insertValueToSheet(self.sheet_KM332A, '2线制', 1, ['3', '4', '5'])
            self.insertValueToSheet(self.sheet_KM332A, '2线制', 2, ['23', '24', '25'])
            self.insertValueToSheet(self.sheet_KM332A, '2线制', 3, ['8', '9', '10'])
            self.insertValueToSheet(self.sheet_KM332A, '2线制', 4, ['28', '29', '30'])
            self.insertValueToSheet(self.sheet_KM332A, '2线制', 5, ['13', '14', '15'])
            self.insertValueToSheet(self.sheet_KM332A, '2线制', 6, ['33', '34', '35'])
            self.insertValueToSheet(self.sheet_KM332A, '2线制', 7, ['18', '19', '20'])
            self.insertValueToSheet(self.sheet_KM332A, '2线制', 8, ['38', '39', '40'])
            self.insertValueToSheet(self.sheet_KM332A, '4线制', 1, ['3', '4', '5'])
            self.insertValueToSheet(self.sheet_KM332A, '4线制', 2, ['23', '24', '25'])
            self.insertValueToSheet(self.sheet_KM332A, '4线制', 3, ['8', '9', '10'])
            self.insertValueToSheet(self.sheet_KM332A, '4线制', 4, ['28', '29', '30'])
            self.insertValueToSheet(self.sheet_KM332A, '4线制', 5, ['13', '14', '15'])
            self.insertValueToSheet(self.sheet_KM332A, '4线制', 6, ['33', '34', '35'])
            self.insertValueToSheet(self.sheet_KM332A, '4线制', 7, ['18', '19', '20'])
            self.insertValueToSheet(self.sheet_KM332A, '4线制', 8, ['38', '39', '40'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 21, 'KM333A', 'TC', 'False', '8', ['2线制'])
            self.insertValueToSheet(self.sheet_KM333A, '2线制', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM333A, '2线制', 2, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM333A, '2线制', 3, ['6', '7'])
            self.insertValueToSheet(self.sheet_KM333A, '2线制', 4, ['28', '29'])
            self.insertValueToSheet(self.sheet_KM333A, '2线制', 5, ['10', '11'])
            self.insertValueToSheet(self.sheet_KM333A, '2线制', 6, ['32', '33'])
            self.insertValueToSheet(self.sheet_KM333A, '2线制', 7, ['14', '15'])
            self.insertValueToSheet(self.sheet_KM333A, '2线制', 8, ['36', '37'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 22, 'KM333B', 'TC', 'False', '16', ['2线制'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 1, ['1', '2'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 2, ['21', '22'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 3, ['3', '4'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 4, ['23', '24'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 5, ['5', '6'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 6, ['25', '26'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 7, ['14', '15'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 8, ['36', '37'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 9, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 10, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 11, ['6', '7'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 12, ['28', '29'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 13, ['10', '11'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 14, ['32', '33'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 15, ['14', '15'])
            self.insertValueToSheet(self.sheet_KM333B, '2线制', 16, ['36', '37'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 23, 'KM334A', 'DI', 'False', '16', ['2线制'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 1, ['2', '22'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 2, ['3', '23'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 3, ['4', '24'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 4, ['5', '25'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 5, ['7', '27'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 6, ['8', '28'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 7, ['9', '29'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 8, ['10', '30'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 9, ['12', '32'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 10, ['13', '33'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 11, ['14', '34'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 12, ['15', '35'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 13, ['17', '37'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 14, ['18', '38'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 15, ['19', '39'])
            self.insertValueToSheet(self.sheet_KM334A, '2线制', 16, ['20', '40'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 24, 'KM335A', 'DO', 'False', '8', ['2线制'])
            self.insertValueToSheet(self.sheet_KM335A, '2线制', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM335A, '2线制', 2, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM335A, '2线制', 3, ['7', '8'])
            self.insertValueToSheet(self.sheet_KM335A, '2线制', 4, ['29', '30'])
            self.insertValueToSheet(self.sheet_KM335A, '2线制', 5, ['12', '13'])
            self.insertValueToSheet(self.sheet_KM335A, '2线制', 6, ['34', '35'])
            self.insertValueToSheet(self.sheet_KM335A, '2线制', 7, ['17', '18'])
            self.insertValueToSheet(self.sheet_KM335A, '2线制', 8, ['39', '40'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 25, 'KM335B', 'DO', 'False', '16', ['2线制'])
            self.insertValueToSheet(self.sheet_KM335B, '2线制', 1, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM335B, '2线制', 2, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM335B, '2线制', 3, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM335B, '2线制', 4, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM335B, '2线制', 5, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM335B, '2线制', 6, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM335B, '2线制', 7, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM335B, '2线制', 8, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM335B, '2线制', 9, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM335B, '2线制', 10, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM335B, '2线制', 12, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM335B, '2线制', 13, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM335B, '2线制', 14, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM335B, '2线制', 15, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM335B, '2线制', 16, ['转接', '至', '组件'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 26, 'KM336A', 'AO', 'False', '6', ['2线制'])
            self.insertValueToSheet(self.sheet_KM336A, '2线制', 1, ['2', '3'])
            self.insertValueToSheet(self.sheet_KM336A, '2线制', 2, ['24', '25'])
            self.insertValueToSheet(self.sheet_KM336A, '2线制', 3, ['7', '8'])
            self.insertValueToSheet(self.sheet_KM336A, '2线制', 4, ['29', '30'])
            self.insertValueToSheet(self.sheet_KM336A, '2线制', 5, ['12', '13'])
            self.insertValueToSheet(self.sheet_KM336A, '2线制', 6, ['34', '35'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 27, 'KM531A', 'LVDT', 'False', '1', ['3线制'])
            self.insertValueToSheet(self.sheet_KM531A, '3线制', 1, ['3', '4', '6'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 28, 'KM532A', 'LVDT', 'False', '1', ['3线制'])
            self.insertValueToSheet(self.sheet_KM532A, '3线制', 1, ['3', '4', '6'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 29, 'KM532B', 'LVDT', 'False', '2', ['3线制'])
            self.insertValueToSheet(self.sheet_KM532B, '3线制', 1, ['3', '4', '6'])
            self.insertValueToSheet(self.sheet_KM532B, '3线制', 2, ['23', '24', '26'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 30, 'KM533A', 'LVDT', 'False', '3', ['L', 'H'])
            self.insertValueToSheet(self.sheet_KM533A, 'L', 1, ['18', '19', '20'])
            self.insertValueToSheet(self.sheet_KM533A, 'L', 2, ['34', '35', '36'])
            self.insertValueToSheet(self.sheet_KM533A, 'L', 3, ['3(V)', 'None', '4(C)'])
            self.insertValueToSheet(self.sheet_KM533A, 'H', 1, ['10', '11', '12'])
            self.insertValueToSheet(self.sheet_KM533A, 'H', 2, ['26', '27', '28'])
            self.insertValueToSheet(self.sheet_KM533A, 'H', 3, ['3(V)', 'None', '4(C)'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 31, 'KM534A', 'LVDT', 'False', '4', ['4线制'])
            self.insertValueToSheet(self.sheet_KM534A, '4线制', 1, ['2', '3', '4'])
            self.insertValueToSheet(self.sheet_KM534A, '4线制', 2, ['27', '28', '29'])
            self.insertValueToSheet(self.sheet_KM534A, '4线制', 3, ['12', '13', '14'])
            self.insertValueToSheet(self.sheet_KM534A, '4线制', 4, ['37', '38', '39'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 32, 'KM535A', 'LVDT', 'False', '4', ['4线制'])
            self.insertValueToSheet(self.sheet_KM535A, '4线制', 1, ['None', '3', '4'])
            self.insertValueToSheet(self.sheet_KM535A, '4线制', 2, ['None', '27', '28'])
            self.insertValueToSheet(self.sheet_KM535A, '4线制', 3, ['None', '13', '14'])
            self.insertValueToSheet(self.sheet_KM535A, '4线制', 4, ['None', '37', '38'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 33, 'KM536A', 'LVDT', 'False', '2', ['4线制'])
            self.insertValueToSheet(self.sheet_KM536A, '4线制', 1, ['26(A1)', '27(B1)', '28(G1)'])
            self.insertValueToSheet(self.sheet_KM536A, '4线制', 2, ['33(A2)', '34(B2)', '35(G2)'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 34, 'KM631A', 'RTU', 'True', '2', ['4线制'])
            self.insertValueToSheet(self.sheet_KM631A, '4线制', 1, ['26(A1)', '27(B1)', '28(G1)'])
            self.insertValueToSheet(self.sheet_KM631A, '4线制', 2, ['33(A2)', '34(B2)', '35(G2)'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 35, 'KM631B', 'RTU', 'True', '2', ['2线制'])
            self.insertValueToSheet(self.sheet_KM631B, '4线制', 1, ['24(R+)', '25(T+)'])
            self.insertValueToSheet(self.sheet_KM631B, '4线制', 2, ['5(R-)', '6(T-)'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 36, 'KM631C', 'RTU', 'True', '2', ['2线制', '4线制'])
            self.insertValueToSheet(self.sheet_KM631C, '2线制', 1, ['24(R+)', '25(T+)'])
            self.insertValueToSheet(self.sheet_KM631C, '2线制', 2, ['5(R-)', '6(T-)'])
            self.insertValueToSheet(self.sheet_KM631C, '4线制', 1, ['33(A)', '34(B)'])
            self.insertValueToSheet(self.sheet_KM631C, '4线制', 2, [])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 37, 'KM632A', 'DP', 'False', '2', ['4线制'])
            self.insertValueToSheet(self.sheet_KM632A, '4线制', 1, ['27(A1)', '28(B1)', '29(G1)'])
            self.insertValueToSheet(self.sheet_KM632A, '4线制', 2, ['34(A2)', '35(B2)', '36(G2)'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 38, 'KM632C', 'DP', 'False', '2', ['2线制'])
            self.insertValueToSheet(self.sheet_KM632C, '2线制', 1, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM632C, '2线制', 2, ['转接', '至', '组件'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 39, 'KM633A', 'PA', 'False', '2', ['2线制'])
            self.insertValueToSheet(self.sheet_KM633A, '2线制', 1, ['18', '19', '20'])
            self.insertValueToSheet(self.sheet_KM633A, '2线制', 2, ['38', '39', '40'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 40, 'KM633B', 'DP/PA', 'True', '2', ['2线制'])
            self.insertValueToSheet(self.sheet_KM633B, '2线制', 1, ['转接', '至', '组件'])
            self.insertValueToSheet(self.sheet_KM633B, '2线制', 2, ['转接', '至', '组件'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 41, 'KB431A', 'DQ', 'False', '6', ['2线制'])
            self.insertValueToSheet(self.sheet_KB431A, '2线制', 1, ['1', '2'])
            self.insertValueToSheet(self.sheet_KB431A, '2线制', 2, ['9', '10'])
            self.insertValueToSheet(self.sheet_KB431A, '2线制', 3, ['17', '18'])
            self.insertValueToSheet(self.sheet_KB431A, '2线制', 4, ['1', '2'])
            self.insertValueToSheet(self.sheet_KB431A, '2线制', 5, ['9', '10'])
            self.insertValueToSheet(self.sheet_KB431A, '2线制', 6, ['17', '18'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 42, 'KB432A', 'DQ', 'True', '16', ['2线制'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 1, ['1', '2'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 2, ['4', '5'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 3, ['6', '7'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 4, ['9', '10'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 5, ['11', '12'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 6, ['14', '15'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 7, ['16', '17'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 8, ['19', '20'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 9, ['1', '2'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 10, ['4', '5'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 11, ['6', '7'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 12, ['9', '10'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 13, ['11', '12'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 14, ['14', '15'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 15, ['16', '17'])
            self.insertValueToSheet(self.sheet_KB432A, '2线制', 16, ['19', '20'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 43, 'KB432C', 'DQ', 'False', '14', ['2线制'])
            self.insertValueToSheet(self.sheet_KB432C, '2线制', 1, ['1', '2'])
            self.insertValueToSheet(self.sheet_KB432C, '2线制', 2, ['4', '5'])
            self.insertValueToSheet(self.sheet_KB432C, '2线制', 3, ['6', '7'])
            self.insertValueToSheet(self.sheet_KB432C, '2线制', 4, ['9', '10'])
            self.insertValueToSheet(self.sheet_KB432C, '2线制', 5, ['11', '12'])
            self.insertValueToSheet(self.sheet_KB432C, '2线制', 6, ['14', '15'])
            self.insertValueToSheet(self.sheet_KB432C, '2线制', 7, ['16', '17'])
            self.insertValueToSheet(self.sheet_KB432C, '2线制', 8, ['19', '20'])
            self.insertValueToSheet(self.sheet_KB432C, '2线制', 9, ['1', '2'])
            self.insertValueToSheet(self.sheet_KB432C, '2线制', 10, ['4', '6'])
            self.insertValueToSheet(self.sheet_KB432C, '2线制', 11, ['6', '7'])
            self.insertValueToSheet(self.sheet_KB432C, '2线制', 12, ['9', '10'])
            self.insertValueToSheet(self.sheet_KB432C, '2线制', 13, ['12', '14'])
            self.insertValueToSheet(self.sheet_KB432C, '2线制', 14, ['16', '18'])
            self.insertValueToCartonSheet(self.sheet_fixedCarton, 44, 'KB432D', 'DQ', 'True', '16', ['2线制'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 1, ['5', '6'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 2, ['7', '8'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 3, ['9', '10'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 4, ['11', '12'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 5, ['13', '14'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 6, ['15', '16'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 7, ['17', '18'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 8, ['19', '20'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 9, ['5', '6'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 10, ['7', '8'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 11, ['9', '10'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 12, ['11', '12'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 13, ['13', '14'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 14, ['15', '16'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 15, ['17', '18'])
            self.insertValueToSheet(self.sheet_KB432D, '2线制', 16, ['19', '20'])

            self.wb_carton.save(self.basePath)

    def insertValueToCartonSheet(self, sheet, row, value_carton, value_nameMin, value_suport, value_roadNum, value_wires):
        if row == 2:
            if sheet['A1'].value is None or sheet['A1'].value == 'None' or sheet['A1'].value == '':
                row = 1

        sheet.cell(row, 1, value_carton)
        sheet.cell(row, 2, value_nameMin)
        sheet.cell(row, 3, value_suport)
        sheet.cell(row, 4, value_roadNum)

        num_wireSuport = len(value_wires)

        for n in range(0, num_wireSuport):
            sheet.cell(row, 5+n, value_wires[n])

    def insertValueToSheet(self, sheet, str_wire, num_road, thoroughfare):
        if str_wire == '2线制':
            index_childSheetForm = 0
        elif str_wire == '3线制':
            index_childSheetForm = 1
        elif str_wire == '4线制':
            index_childSheetForm = 2
        elif str_wire == '220V':
            index_childSheetForm = 3
        elif str_wire == 'OC':
            index_childSheetForm = 4
        elif str_wire == 'TTL':
            index_childSheetForm = 5
        elif str_wire == 'L':
            index_childSheetForm = 6
        elif str_wire == 'H':
            index_childSheetForm = 7
        else:
            ...

        if len(thoroughfare) == 0:
            thoroughfare = ['None', 'None', 'None', 'None']
        elif len(thoroughfare) == 1:
            thoroughfare = thoroughfare + ['None', 'None', 'None']
        elif len(thoroughfare) == 2:
            thoroughfare = thoroughfare + ['None', 'None']
        elif len(thoroughfare) == 3:
            thoroughfare = thoroughfare + ['None']


        sheet.cell(1 + index_childSheetForm * 20, 1, str_wire)
        sheet.cell(2 + num_road - 1 + index_childSheetForm * 20, 2, thoroughfare[0])
        sheet.cell(2 + num_road - 1 + index_childSheetForm * 20, 3, thoroughfare[1])
        sheet.cell(2 + num_road - 1 + index_childSheetForm * 20, 4, thoroughfare[2])
        sheet.cell(2 + num_road - 1 + index_childSheetForm * 20, 5, thoroughfare[3])

    def insertValueToCustomizeCartonSheet(self, dict_carton_confirmed):
        """
        dict_carton_confirmed = {
            'name_carton': '',
            'nameMin': '',
            'suport': '',
            'numRoad': 0,
            # 以上均为字符串
            'wires': [],
            'thoroughfare_wire': []  # 内容为各通道接线，数量等同于通道数
        }
        """
        cartons_exited = self.getCartons('通用')
        if dict_carton_confirmed['name_carton'] in cartons_exited:
            return 0
        wb = openpyxl.load_workbook(self.basePath)
        sheet = wb['卡件汇总(自定义)']
        num_sheets = len(wb.sheetnames)
        row = sheet.max_row
        self.insertValueToCartonSheet(sheet, row+1, dict_carton_confirmed['name_carton'],
                                      dict_carton_confirmed['nameMin'], dict_carton_confirmed['suport'],
                                      str(dict_carton_confirmed['numRoad']), dict_carton_confirmed['wires'])

        sheet_currentCarton = wb.create_sheet(dict_carton_confirmed['name_carton'], num_sheets)

        for wire in dict_carton_confirmed['wires']:
            thoroughfare_oneWire = dict_carton_confirmed['thoroughfare_wire'][dict_carton_confirmed['wires'].index(wire)]
            for n in range(0, int(dict_carton_confirmed['numRoad'])):
                self.insertValueToSheet(sheet_currentCarton, wire, n+1, thoroughfare_oneWire[n])

        wb.save(self.basePath)
        wb.close()

        return 1

    def getCartons(self, cabinet):
        cartons_fixed = []
        cartons_confirmed = []
        wb_carton = openpyxl.load_workbook(self.basePath)
        sheet_fixed = wb_carton['卡件汇总(固定)']
        sheet_confirmed = wb_carton['卡件汇总(自定义)']
        cartons_fixed_cell = sheet_fixed["A"]
        cartons_confirmed_cell = sheet_confirmed["A"]

        if cabinet == '通用':
            for cell_carton in cartons_fixed_cell:
                cartons_fixed.append(cell_carton.value)
        elif cabinet == '高密度':
            for cell_carton in cartons_fixed_cell:
                cell_super = sheet_fixed.cell(row=cell_carton.row, column=3)
                if cell_super.value == 'True' or cell_super.value == True:
                    cartons_fixed.append(cell_carton.value)
                elif cell_super.value == 'False' or cell_super.value == False:
                    continue
        else:
            ...

        if sheet_confirmed['A1'].value is None or sheet_confirmed['A1'].value == 'None' or sheet_confirmed[
            'A1'].value == '':
            cartons = cartons_fixed

            wb_carton.save(self.basePath)
            wb_carton.close()
            return cartons

        if cabinet == '通用':
            for cell_carton in cartons_confirmed_cell:
                cartons_confirmed.append(cell_carton.value)
        elif cabinet == '高密度':
            for cell_carton in cartons_confirmed_cell:
                cell_super = sheet_confirmed.cell(row=cell_carton.row, column=3)
                if cell_super.value == 'True' or cell_super.value == True:
                    cartons_confirmed.append(cell_carton.value)
                elif cell_super.value == 'False' or cell_super.value == False:
                    continue
        else:
            ...

        if cartons_confirmed == []:
            cartons = cartons_fixed
        else:
            cartons = cartons_fixed + cartons_confirmed

        wb_carton.save(self.basePath)
        wb_carton.close()
        return cartons

    def getWires(self, str_carton):

        wires = []

        wb_carton = openpyxl.load_workbook(self.basePath)
        sheet_fixed = wb_carton['卡件汇总(固定)']
        cartons_cell = sheet_fixed["A"]

        for cell_carton in cartons_cell:
            if cell_carton.value == str_carton:
                values_line = sheet_fixed[cell_carton.row]
            else:
                continue

        num_line = len(values_line)
        for i in range(4, num_line):
            wire = values_line[i].value
            if wire == '' or wire == 'None' or wire == None:
                continue
            else:
                wires.append(wire)

        wb_carton.save(self.basePath)
        wb_carton.close()

        return wires

    def getCartonInfo(self, str_carton, str_need):
        wb_carton = openpyxl.load_workbook(self.basePath)
        sheet_fixed = wb_carton['卡件汇总(固定)']
        cartons_cell = sheet_fixed["A"]

        for cell_carton in cartons_cell:
            if cell_carton.value == str_carton:
                values_line = sheet_fixed[cell_carton.row]
            else:
                continue

        if str_need == 'nameMin':
            index = 1
        elif str_need == 'roadNum':
            index = 3
        elif str_need == 'wire_first':
            index = 4

        value_need = values_line[index].value

        wb_carton.save(self.basePath)
        wb_carton.close()

        return value_need

    def getThoroughfare(self, str_carton, num_road, str_wire):
        throughconnect = []

        wb_carton = openpyxl.load_workbook(self.basePath)
        sheet = wb_carton[str_carton]

        if str_wire == '2线制':
            index_childSheetForm = 0
        elif str_wire == '3线制':
            index_childSheetForm = 1
        elif str_wire == '4线制':
            index_childSheetForm = 2
        elif str_wire == '220V':
            index_childSheetForm = 3
        elif str_wire == 'OC':
            index_childSheetForm = 4
        elif str_wire == 'TTL':
            index_childSheetForm = 5
        elif str_wire == 'L':
            index_childSheetForm = 6
        elif str_wire == 'H':
            index_childSheetForm = 7
        else:
            ...

        cell1 = sheet.cell(2 + int(num_road) - 1 + index_childSheetForm * 20, 2)
        cell2 = sheet.cell(2 + int(num_road) - 1 + index_childSheetForm * 20, 3)
        cell3 = sheet.cell(2 + int(num_road) - 1 + index_childSheetForm * 20, 4)
        cell4 = sheet.cell(2 + int(num_road) - 1 + index_childSheetForm * 20, 5)

        value1 = cell1.value
        value2 = cell2.value
        value3 = cell3.value
        value4 = cell4.value

        if cell1.value == 'None' or cell1.value == '' or cell1.value == None:
            value1 = ''
        if cell2.value == 'None' or cell2.value == '' or cell2.value == None:
            value2 = ''
        if cell3.value == 'None' or cell3.value == '' or cell3.value == None:
            value3 = ''
        if cell4.value == 'None' or cell4.value == '' or cell4.value == None:
            value4 = ''

        throughconnect.append(value1)
        throughconnect.append(value2)
        throughconnect.append(value3)
        throughconnect.append(value4)

        wb_carton.save(self.basePath)
        wb_carton.close()

        return throughconnect

    def getCurrentCabinetCartons(self):
        chooses = self.getCartons('通用')
        return chooses

    def getHighDensityCabinetCartons(self):
        try:
            chooses = self.getCartons('高密度')
        except:
            self.logger.tranceInfo()

        return chooses

    def getTerminalData(self, name_carton):
        list_cartonData = []
        numRoad = int(self.getCartonInfo(name_carton, 'roadNum'))
        wire_first = self.getCartonInfo(name_carton, 'wire_first')
        nameMin = self.getCartonInfo(name_carton, 'nameMin')

        for n in range(0, numRoad):
            list_row = []
            connect = self.getThoroughfare(name_carton, n+1, wire_first)
            list_row.append(name_carton)
            list_row.append(nameMin)
            list_row.append(wire_first)
            list_row.append(str(n+1))
            list_row.append(connect[0])
            list_row.append(connect[1])
            list_row.append(connect[2])
            list_row.append(connect[3])
            list_cartonData.append(list_row)

        return list_cartonData


# 测试函数
def test():
    obj_thisPyFile = PotsDataVessel()

# 文件运行
if __name__ == '__main__':
    test()
